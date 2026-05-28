"""Excel upload parser + template generator (Mode B, decision #19).

Excel schema (decision #19):
  · Manufacture Part Number (required)
  · Manufacture (optional — used as MPN hint to pipeline)
  · Type (optional — webapp-only metadata, decision #29 overlays chip-list join)
  · risk (optional — webapp-only metadata)

Template generation: just a header-only xlsx with the 4 columns + a sample row.
"""
from __future__ import annotations
import io
from pathlib import Path
from typing import BinaryIO

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

REQUIRED_COL = "Manufacture Part Number"
OPTIONAL_COLS = ["Manufacture", "Type", "risk"]
ALL_COLS = [REQUIRED_COL] + OPTIONAL_COLS

# Sample row for the downloadable template
TEMPLATE_SAMPLE_ROW = [
    "STM32G030F6P6",   # required: MPN
    "STMicroelectronics",  # optional: manufacturer hint
    "MCU",             # optional: business type
    "high",            # optional: business risk (high / low)
]


def make_template_xlsx() -> bytes:
    """Generate a header-only Excel template (returned as bytes for download)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MPN List"

    fill = PatternFill(start_color="FF1F4E78", end_color="FF1F4E78", fill_type="solid")
    font_white = Font(bold=True, color="FFFFFFFF", name="Calibri")

    for idx, col in enumerate(ALL_COLS, start=1):
        cell = ws.cell(row=1, column=idx, value=col)
        cell.fill = fill
        cell.font = font_white
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.column_dimensions[cell.column_letter].width = 24

    # Sample row (greyed out via comment-style note in row 2)
    for idx, val in enumerate(TEMPLATE_SAMPLE_ROW, start=1):
        ws.cell(row=2, column=idx, value=val).font = Font(name="Calibri", italic=True, color="FF888888")

    # Hint note in row 3
    ws.cell(row=3, column=1,
            value="↑ 上面是示例行（斜体灰色）。请删除后填入你的数据。"
                  "Manufacture Part Number 必填；其他列可选。"
            ).font = Font(name="Calibri", italic=True, color="FF888888")
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=4)

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class ExcelParseError(ValueError):
    pass


def parse_upload(file_bytes: bytes) -> tuple[list[str], list[dict], int]:
    """Parse an uploaded xlsx → (mpns_for_pipeline, metadata_records, raw_row_count).

    mpns_for_pipeline: list of MPN strings, deduplicated (first occurrence kept).
                       Will be further normalised by mpn_cleaner.
    metadata_records:  list of {Manufacture Part Number, Manufacture, Type, risk},
                       one per dedup'd MPN; for join in T4 per decision #29.
    raw_row_count:     number of MPN-bearing rows in the upload BEFORE dedup
                       (so the review page can show "uploaded N rows → M unique").

    Raises ExcelParseError on malformed input.
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as e:
        raise ExcelParseError(f"无法读取 xlsx 文件：{e}")

    # Use the first sheet (template uses "MPN List", but accept any)
    ws = wb[wb.sheetnames[0]]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = list(next(rows_iter))
    except StopIteration:
        raise ExcelParseError("xlsx 是空的")

    header_clean = [str(h).strip() if h is not None else "" for h in header]

    # Validate required column
    if REQUIRED_COL not in header_clean:
        raise ExcelParseError(
            f"模板格式不对——缺列 '{REQUIRED_COL}'。"
            f"实际列：{header_clean[:8]}。请下载最新模板。"
        )

    col_idx = {name: header_clean.index(name) for name in ALL_COLS if name in header_clean}

    mpns: list[str] = []
    seen_mpns: set[str] = set()
    metadata: list[dict] = []
    raw_row_count = 0

    for row_num, row in enumerate(rows_iter, start=2):
        if row is None or all(c is None for c in row):
            continue
        raw_mpn = row[col_idx[REQUIRED_COL]]
        if raw_mpn is None:
            continue
        mpn = str(raw_mpn).strip()
        if not mpn:
            continue
        # Skip sample row from template
        if mpn == "STM32G030F6P6" and row_num == 2:
            sample_match = True
            for sample_idx, sample_val in enumerate(TEMPLATE_SAMPLE_ROW):
                col = ALL_COLS[sample_idx]
                if col in col_idx:
                    actual = row[col_idx[col]]
                    if actual is not None and str(actual).strip() != sample_val:
                        sample_match = False
                        break
            if sample_match:
                continue

        raw_row_count += 1
        if mpn in seen_mpns:
            continue
        seen_mpns.add(mpn)
        mpns.append(mpn)

        rec = {"Manufacture Part Number": mpn}
        for col in OPTIONAL_COLS:
            if col in col_idx:
                v = row[col_idx[col]]
                rec[col] = str(v).strip() if v is not None else ""
            else:
                rec[col] = ""
        metadata.append(rec)

    if not mpns:
        raise ExcelParseError("xlsx 里没有任何 MPN 数据（除了示例行）")

    return mpns, metadata, raw_row_count


def write_input_csv(metadata: list[dict], out_path: Path) -> None:
    """Save business metadata to webapp/runs/<run_id>/input.csv (decision #29)."""
    import csv
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=ALL_COLS)
        w.writeheader()
        w.writerows(metadata)
