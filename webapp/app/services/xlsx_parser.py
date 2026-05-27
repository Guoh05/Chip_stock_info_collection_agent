"""Read pipeline's merged xlsx → filter + sort → list[dict].

Implements decisions #25 (in_stock=True filter), #26 (custom sort),
#33 (column-name-based dispatch, unknown columns logged & ignored).

Returns TWO lists:
  - `filtered_rows`: passed in_stock filter, sorted per decision #26 — used for
                     parsed.json and web/email rendering
  - `all_rows`:      every row from All_data sheet — used for slim xlsx download
                     (decision #30: download keeps in_stock=False rows for
                      lead-time analysis)
"""
from __future__ import annotations
import logging
from pathlib import Path

import openpyxl

from ..schemas import WEBAPP_SCHEMA_v1

log = logging.getLogger("webapp.parser")

# Decision #26 sort: risk(high→low→other→null) → Type → MPN → Broker → qty desc
RISK_RANK = {"high": 0, "low": 1, "medium": 2, "med": 2, "": 3, None: 3}


def _is_truthy(v) -> bool:
    return v in (True, 1, "True", "true", "TRUE", "1")


def parse_merged_xlsx(xlsx_path: Path) -> tuple[list[dict], list[dict]]:
    """Read All_data sheet from pipeline merged xlsx.

    Returns (filtered_sorted_rows, all_rows).
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    if "All_data" not in wb.sheetnames:
        raise ValueError(f"Sheet 'All_data' missing in {xlsx_path}. Sheets: {wb.sheetnames}")
    ws = wb["All_data"]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = list(next(rows_iter))
    except StopIteration:
        return [], []

    # Schema drift detection (decision #33).
    known = set(WEBAPP_SCHEMA_v1.keys())
    expected_t1_t2 = {k for k, v in WEBAPP_SCHEMA_v1.items() if v["tier"] in (1, 2)}
    actual_cols = {c for c in header if c}
    missing_for_display = expected_t1_t2 - actual_cols
    unknown_to_webapp = actual_cols - known
    # T3 + ref_* and unknown are all "extra"; only warn on cols neither in schema
    # nor obviously ref/audit. Don't spam on ref_*.
    unknown_to_webapp = {c for c in unknown_to_webapp if not c.startswith("ref_")}
    if missing_for_display:
        log.warning(
            "[schema] columns in WEBAPP_SCHEMA_v1 missing from xlsx %s: %s — "
            "those cells will render as '—'",
            xlsx_path.name, sorted(missing_for_display),
        )
    if unknown_to_webapp:
        log.info(
            "[schema] columns in xlsx not in WEBAPP_SCHEMA_v1 (kept in slim xlsx, "
            "ignored for web): %s", sorted(unknown_to_webapp),
        )

    all_rows: list[dict] = []
    for row in rows_iter:
        if row is None or all(c is None for c in row):
            continue
        rec = {col: val for col, val in zip(header, row) if col is not None}
        all_rows.append(rec)

    # Decision #25: in_stock filter for display
    filtered = [r for r in all_rows if _is_truthy(r.get("in_stock"))]

    # Decision #26: sort risk → Type → MPN → Broker → qty desc
    def sort_key(r):
        risk = (str(r.get("risk") or "")).strip().lower()
        return (
            RISK_RANK.get(risk, 3),
            str(r.get("Type") or ""),
            str(r.get("MPN_cleaned_byAgent") or ""),
            str(r.get("Broker name") or ""),
            -(r.get("Available Quantity") or 0),
        )

    filtered.sort(key=sort_key)
    return filtered, all_rows
