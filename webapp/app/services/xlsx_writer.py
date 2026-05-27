"""Generate slim xlsx for download (decision #30).

The pipeline produces a 5-sheet xlsx; webapp serves a 1-sheet variant containing
only `All_data` (all rows, including in_stock=False — kept for lead-time
analysis per decision #25 trade-off).

Visual conventions sync the pipeline (decision #14): 3-zone header palette
(A-L blue, M-AH orange, AI+ grey) + 8 procurement-key dark-red headers
+ Calibri font + in_stock=True light-green row tint + freeze panes + AutoFilter.
"""
from __future__ import annotations
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..schemas import HIGHLIGHT_COLUMNS

HEADER_BLUE = PatternFill(start_color="FF1F4E78", end_color="FF1F4E78", fill_type="solid")
HEADER_ORANGE = PatternFill(start_color="FFFCE4D6", end_color="FFFCE4D6", fill_type="solid")
HEADER_GREY = PatternFill(start_color="FF595959", end_color="FF595959", fill_type="solid")
HEADER_RED = PatternFill(start_color="FFC00000", end_color="FFC00000", fill_type="solid")
FONT_WHITE_BOLD = Font(bold=True, color="FFFFFFFF", name="Calibri")
FONT_BLACK_BOLD = Font(bold=True, color="FF000000", name="Calibri")
FONT_BODY = Font(name="Calibri")
ROW_GREEN = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
ROW_GREY = PatternFill(start_color="FFEEEEEE", end_color="FFEEEEEE", fill_type="solid")


def _header_style(col_idx: int, col_name: str) -> tuple[PatternFill, Font]:
    if col_name in HIGHLIGHT_COLUMNS:
        return HEADER_RED, FONT_WHITE_BOLD
    if col_idx <= 12:
        return HEADER_BLUE, FONT_WHITE_BOLD
    if col_idx <= 34:
        return HEADER_ORANGE, FONT_BLACK_BOLD
    return HEADER_GREY, FONT_WHITE_BOLD


def _row_fill(record: dict) -> PatternFill | None:
    in_stock = record.get("in_stock")
    if in_stock in (True, 1, "True", "true", "TRUE", "1"):
        return ROW_GREEN
    qty = record.get("Available Quantity")
    if in_stock is not None and (qty == 0 or qty == "0"):
        return ROW_GREY
    return None


def write_slim_xlsx(rows: list[dict], out_path: Path, header_order: list[str] | None = None) -> None:
    """Write a single-sheet `All_data` xlsx.

    rows = list of dicts (column name → cell value). `header_order` if given
    fixes column order; otherwise the first row's keys are used.
    """
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "All_data"

    if not rows:
        ws.cell(row=1, column=1, value="(no data — pipeline produced 0 rows)").font = FONT_BODY
        wb.save(out_path)
        return

    headers = header_order or list(rows[0].keys())

    # Headers
    for idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx, value=h)
        fill, font = _header_style(idx, h)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    # Body
    for row_idx, record in enumerate(rows, start=2):
        fill = _row_fill(record)
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=record.get(h))
            cell.font = FONT_BODY
            if fill is not None:
                cell.fill = fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"

    # Light column-width hints (mimic pipeline's COLUMN_WIDTHS philosophy)
    for idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 16

    wb.save(out_path)
