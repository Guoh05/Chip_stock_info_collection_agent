"""Merge latest API + Scraper batch_index.csv into a procurement-facing xlsx.

Rules:
1. status=ok only.
2. Drop source HQEW_*.
3. API wins per (input_mpn, source): if the API CSV has any status=ok rows for
   that pair, drop ALL scraper rows for the same pair (even if API qty=0).
4. mfr_match=False rows are kept; flagged via column.
5. in_stock = stockpool_qty is not None AND > 0.
6. Cross-validation: for (mpn, source) where BOTH tracks have status=ok rows
   and no scraper qty matches any API warehouse qty, scraper rows go to a
   reference sheet (Sheet 3) for QA.
7. Output columns are renamed / reordered per
   ref/merged_output_fields_mapping.xlsx and enriched with project metadata
   from ref/Raw_chip_list_20260520.xlsx (sheet "Part List
   Modify"). Lead time converted from days to weeks (1 decimal).
8. Sheet 1 ("High_risk_positive_stock") = procurement priority view:
   risk == "high" AND Available Quantity > 0.
9. v1.10 — `Manufacture Part Number` column now shows the RAW MPN from the
   chip list; new `MPN_cleaned_byAgent` column shows the agent-cleaned MPN
   actually sent to sources. Chip-list join uses `MPN_cleaned` column when
   present, falling back to `Manufacture Part Number` for legacy chip lists.
   See CLAUDE.md Hard Rule #8 + memory `feedback_input_review.md`.

Output: <env_root>/merged/Merge_<api_ts>__<scr_ts>/
        Versuni_chip_stock_availability_check_<YYYYMMDD>.xlsx
        + Versuni_chip_stock_availability_check_<YYYYMMDD>.csv (= Sheet 2)
        (<YYYYMMDD> is the date the merge was executed.)

`<env_root>` is `test/` (default) or `production/` (with --env prod). The
same flag also picks where to read the API + Scraper batches from
(`<env_root>/api/` and `<env_root>/scraper/`).
"""
from __future__ import annotations

import argparse
import csv
import re
import sys
from collections import defaultdict
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parents[1]
ENV_ROOTS = {
    "test": PROJECT_ROOT / "test",
    "prod": PROJECT_ROOT / "production",
}
CHIP_LIST_PATH = PROJECT_ROOT / "ref" / "Raw_chip_list_20260520.xlsx"
CHIP_LIST_SHEET = "Part List Modify"

DROP_SOURCE_PREFIXES = ("HQEW_",)

# Chip-list headers we copy into the merged output (first-row per MPN).
# v1.10 — `Manufacture Part Number` is included here so the output can
# display the raw (un-cleaned) MPN from the chip list alongside the cleaned
# `MPN_cleaned_byAgent` column. See CLAUDE.md Hard Rule #8 for the cleaning
# convention and load_chip_meta() for the dual-key lookup behavior.
CHIP_FIELDS = [
    "Category", "Project", "EMS/Finish Goods", "12NC_PCBA",
    "Manufacture Part Number",
    "Quantity", "Currency", "Current Price", "Type", "risk",
]

# Per-column source:
#   ("merge", <old_csv_key>)        → copy from upstream CSV row (renamed)
#   ("chip",  <chip_list_header>)   → look up in chip_meta[normalized_mpn]
#   ("lead_time", <old_csv_key>)    → days→weeks transform
#   ("computed", None)              → filled in post-processing (Is_orig_manufacture,
#                                      Is_cheapest, price_rank). build_output_row
#                                      leaves these as None.
#   ("blank", None)                 → empty (business to fill)
COLUMN_SOURCE_MAP: dict[str, tuple[str, str | None]] = {
    "Category":                              ("chip", "Category"),
    "Project":                               ("chip", "Project"),
    "EMS/Finish Goods":                      ("chip", "EMS/Finish Goods"),
    "12NC_PCBA":                             ("chip", "12NC_PCBA"),
    "Manufacture Part Number":               ("chip", "Manufacture Part Number"),  # v1.10: raw MPN from chip list (was input_mpn); fallback to input_mpn when no chip-list match
    "MPN_cleaned_byAgent":                   ("merge", "input_mpn"),                # v1.10: agent-cleaned MPN actually sent to sources — traceability column per CLAUDE.md Hard Rule #8
    "Manufacture":                           ("merge", "expected_mfr"),
    "Quantity":                              ("chip", "Quantity"),
    "Currency":                              ("chip", "Currency"),
    "Current Price":                         ("chip", "Current Price"),
    "Type":                                  ("chip", "Type"),
    "risk":                                  ("chip", "risk"),
    "in_stock":                              ("merge", "in_stock"),         # v1.9 A: moved here (after risk)
    "Broker name":                           ("merge", "source"),
    "Data collect method":                   ("merge", "track"),
    "Warehouse/vender":                      ("merge", "warehouse"),
    "Is_orig_manufacture":                   ("computed", None),            # v1.9 E
    "Is_cheapest":                           ("computed", None),            # v1.9 F
    "price_rank":                            ("computed", None),            # v1.9 F
    "Stock Location":                        ("merge", "ships_from"),
    "Available Quantity":                    ("merge", "stockpool_qty"),
    "ship infor after order placed":         ("merge", "ship_text"),
    "Lead Time (Week)":                      ("lead_time", "lead_time_days"),
    "MOQ":                                   ("merge", "moq"),
    "Maximum order qty":                     ("merge", "max_break_qty"),    # v1.9 D: moved before min
    "Unit price w/o VAT (max qty)":          ("merge", "price_at_max_qty"), # v1.9 B+C: rename + VAT-strip if CNY
    "Minimum order qty":                     ("merge", "min_break_qty"),
    "Unit price w/o VAT (min qty)":          ("merge", "price_at_min_qty"), # v1.9 B+C
    "Number of price tiers":                 ("merge", "num_price_tiers"),
    "Trade Currency":                        ("merge", "currency"),         # v1.9: dropped literal \n in header
    "Date of Code":                          ("blank", None),
    "Reel/Cut Reel":                         ("blank", None),
    "Certificate of Conformity(Yes/No)":     ("blank", None),
    "ref_Warehouse/vender ID":               ("merge", "warehouse_idx"),
    "ref_returned_mpn":                      ("merge", "returned_mpn"),
    "ref_vendor_sku":                        ("merge", "vendor_sku"),
    "ref_returned_mfr":                      ("merge", "returned_mfr"),
    "ref_mfr_match":                         ("merge", "mfr_match"),
    "ref_is_mirror":                         ("merge", "is_mirror"),
    "ref_datasheet_url":                     ("merge", "datasheet_url"),
    "ref_status":                            ("merge", "status"),
    "ref_error":                             ("merge", "error"),
}
OUTPUT_COLUMNS = list(COLUMN_SOURCE_MAP.keys())   # 41 columns, in output order

# Sheet 3 (cross-validation reference) — narrower subset, same renames, plus
# the unmapped `note` column (not in COLUMN_SOURCE_MAP — handled in builder).
# Per v1.9 decision: Sheet 3 inherits the price rename + VAT strip but NOT the
# computed columns (Is_orig_manufacture / Is_cheapest / price_rank) — meaningless
# on the QA-only mismatched subset.
MISMATCH_COLUMNS = [
    "Category", "Project", "EMS/Finish Goods", "12NC_PCBA",
    "Manufacture Part Number", "MPN_cleaned_byAgent", "Manufacture",
    "Quantity", "Currency", "Current Price", "Type", "risk",
    "Broker name", "in_stock",
    "ref_returned_mpn", "Warehouse/vender", "Available Quantity",
    "ship infor after order placed", "Lead Time (Week)",
    "MOQ", "Unit price w/o VAT (min qty)", "Trade Currency",
    "ref_datasheet_url", "note",
]

LEAD_TIME_HEADER = "Lead Time (Week)"
QTY_HEADER = "Available Quantity"
MAX_PRICE_HEADER = "Unit price w/o VAT (max qty)"
MIN_PRICE_HEADER = "Unit price w/o VAT (min qty)"
TRADE_CURRENCY_HEADER = "Trade Currency"

# v1.9 C — pre-VAT conversion for CNY/RMB rows.
VAT_DIVISOR = 1.13
CNY_LIKE_CURRENCIES = {"CNY", "RMB", "¥"}

# v1.9 G — columns hidden by default in every sheet that carries them.
HIDDEN_COLUMNS_EXPLICIT = {
    "Minimum order qty",
    "Unit price w/o VAT (min qty)",
    "Number of price tiers",
    "price_rank",
}
# (All columns starting with `ref_` are also hidden — handled inline.)

# v1.9 H — sheets hidden by default (still present in the workbook, can be
# unhidden manually in Excel). Sheet 2 stays as the default active.
HIDDEN_SHEETS = {"High_risk_positive_stock", "ref_scraper_api_diff"}

# Per-column widths from ref/merged_header_example (sheet `header`) — matched
# to procurement's preferred layout. Columns not listed fall back to auto-fit.
COLUMN_WIDTHS: dict[str, float] = {
    "Category": 12.0, "Project": 11.0, "EMS/Finish Goods": 18.0, "12NC_PCBA": 27.0,
    "Manufacture Part Number": 25.0, "MPN_cleaned_byAgent": 22.0, "Manufacture": 13.0,
    "Quantity": 10.0, "Currency": 13.0, "Current Price": 15.0, "Type": 10.0, "risk": 13.0,
    "in_stock": 10.0,
    "Broker name": 33.0, "Data collect method": 21.0,
    "Warehouse/vender": 58.9,
    "Is_orig_manufacture": 19.0, "Is_cheapest": 12.0, "price_rank": 11.0,
    "Stock Location": 26.0, "Available Quantity": 29.9,
    "ship infor after order placed": 45.0, "Lead Time (Week)": 18.0,
    "MOQ": 10.0,
    "Maximum order qty": 19.0, "Unit price w/o VAT (max qty)": 26.0,
    "Minimum order qty": 19.0, "Unit price w/o VAT (min qty)": 26.0,
    "Number of price tiers": 23.0,
    "Trade Currency": 15.0,
    "Date of Code": 14.0, "Reel/Cut Reel": 15.0, "Certificate of Conformity(Yes/No)": 35.0,
    "ref_Warehouse/vender ID": 25.0, "ref_returned_mpn": 18.0, "ref_vendor_sku": 19.0,
    "ref_returned_mfr": 31.0, "ref_mfr_match": 15.0, "ref_is_mirror": 13.0,
    "ref_datasheet_url": 50.0, "ref_status": 12.0, "ref_error": 11.0,
    "note": 40.0,
}

# Sort priority for `risk` column. Higher priority = lower rank number.
RISK_RANK = {"high": 0, "low": 1}

# Per-row highlight fills (Sheet 2 only).
GREEN = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
GREY = PatternFill(start_color="FFEEEEEE", end_color="FFEEEEEE", fill_type="solid")

# Three-zone header palette (v1.10 column layout, 42 cols — was 41 in v1.9):
#   A–L   (cols  1–12): chip-list metadata + raw MPN + cleaned MPN + Manufacture..risk → dark blue + white text
#   M–AG  (cols 13–33): per-row distributor data, computed cols, business-fill         → light orange + black text
#   AH+   (cols 34+ ):  ref_* technical / audit fields                                 → dark grey + white text
HEADER_FILL_BLUE = PatternFill(start_color="FF1F4E78", end_color="FF1F4E78", fill_type="solid")
HEADER_FILL_ORANGE = PatternFill(start_color="FFFCE4D6", end_color="FFFCE4D6", fill_type="solid")
HEADER_FILL_GREY = PatternFill(start_color="FF595959", end_color="FF595959", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, color="FFFFFFFF", name="Calibri")
HEADER_FONT_BLACK = Font(bold=True, color="FF000000", name="Calibri")

# v1.11 — 8 procurement-key columns get a dark-red header to draw the eye.
# Overrides the zone-based palette above when matched (by column NAME, not
# index — survives future column reorderings).
HEADER_FILL_DARK_RED = PatternFill(start_color="FFC00000", end_color="FFC00000", fill_type="solid")
HIGHLIGHT_HEADER_COLUMNS: set[str] = {
    "in_stock",
    "Broker name",
    "Warehouse/vender",
    "Is_orig_manufacture",
    "Is_cheapest",
    "Available Quantity",
    "ship infor after order placed",
    "Unit price w/o VAT (max qty)",
}


def _header_style(col_idx: int, col_name: str | None = None) -> tuple[PatternFill, Font]:
    """Return (fill, font) for the header cell.

    v1.11 — when `col_name` is in HIGHLIGHT_HEADER_COLUMNS, override the
    zone-based palette with dark-red + white. Otherwise fall through to the
    3-zone layout: blue (cols 1-12), orange (13-33), grey (34+).
    """
    if col_name in HIGHLIGHT_HEADER_COLUMNS:
        return HEADER_FILL_DARK_RED, HEADER_FONT_WHITE
    if col_idx <= 12:
        return HEADER_FILL_BLUE, HEADER_FONT_WHITE
    if col_idx <= 33:
        return HEADER_FILL_ORANGE, HEADER_FONT_BLACK
    return HEADER_FILL_GREY, HEADER_FONT_WHITE


# ---------- helpers ---------------------------------------------------------

_BATCH_NAME_RE = re.compile(r"^BatchTest_\d{8}_\d{2}_\d{2}_\d{2}$")


def latest_batch(parent: Path) -> Path:
    """Newest folder matching the standard BatchTest_<YYYYMMDD>_<HH>_<MM>_<SS>/
    naming. Ad-hoc probe folders with extra suffixes (e.g.
    `BatchTest_..._bom2buy`) are skipped — their CSV schemas can differ."""
    candidates = sorted(
        (p for p in parent.iterdir() if p.is_dir() and _BATCH_NAME_RE.match(p.name)),
        key=lambda p: p.name,
    )
    if not candidates:
        raise SystemExit(f"No standard BatchTest_<ts>/ folder in {parent}")
    return candidates[-1]


def normalize_mpn(s) -> str:
    if s is None:
        return ""
    return str(s).strip().replace("\xa0", "")


def parse_qty(s):
    if s is None:
        return None
    s = str(s).strip()
    if s == "":
        return None
    try:
        return int(s)
    except ValueError:
        return None


def _normalize_alnum(s) -> str:
    """Strip everything except letters and digits, upper-case the result.
    Used to compare MPNs across punctuation noise: 'BAV99,215' == 'BAV99-215'."""
    return re.sub(r"[^A-Za-z0-9]", "", str(s or "")).upper()


def returned_mpn_matches_input(raw_row: dict) -> bool:
    """True iff input_mpn and returned_mpn agree after stripping punctuation.

    Drops rows where the channel returned a different MPN string than the one
    we searched for — e.g. searched `GD32E230K4T6`, channel returned
    `GD32E230K6T6`. Empty returned_mpn → False (always dropped).
    """
    inp = _normalize_alnum(raw_row.get("input_mpn"))
    ret = _normalize_alnum(raw_row.get("returned_mpn"))
    return bool(inp) and inp == ret


def days_to_weeks(s):
    """Convert a days-as-string value to weeks (1 decimal). None on missing."""
    if s is None:
        return None
    s = str(s).strip()
    if s == "":
        return None
    try:
        return round(int(s) / 7, 1)
    except ValueError:
        try:
            return round(float(s) / 7, 1)
        except ValueError:
            return None


def load_chip_meta(path: Path) -> dict[str, dict]:
    """Build {normalized_mpn: {chip-field: value, ...}} using first-row per MPN.

    Reads sheet `Part List Modify`. Required columns: `Manufacture Part Number`
    plus everything in CHIP_FIELDS. Duplicates after the first are ignored.

    v1.10 — dual-key lookup. The dict key (used to join against the upstream
    `input_mpn`) is the normalized value of:
      - `MPN_cleaned` column if present (chip list was produced by the agent
        cleaner per CLAUDE.md Hard Rule #8), OR
      - `Manufacture Part Number` column as fallback (legacy chip list).
    The raw `Manufacture Part Number` value is always stored in the meta dict
    so the output can show it alongside the cleaned MPN.
    """
    if not path.exists():
        raise SystemExit(f"chip list not found: {path}")
    wb = openpyxl.load_workbook(path, data_only=True)
    if CHIP_LIST_SHEET not in wb.sheetnames:
        raise SystemExit(f"chip list missing sheet '{CHIP_LIST_SHEET}' in {path}")
    ws = wb[CHIP_LIST_SHEET]
    header_to_col: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v is None:
            continue
        header_to_col[str(v).strip()] = c
    needed = ["Manufacture Part Number", *CHIP_FIELDS]
    # `Manufacture Part Number` is already in CHIP_FIELDS as of v1.10 — dedup before
    # checking required-columns presence to avoid spurious "missing" report.
    needed = list(dict.fromkeys(needed))
    missing = [h for h in needed if h not in header_to_col]
    if missing:
        raise SystemExit(f"chip list missing required columns: {missing}")
    # v1.10 — prefer MPN_cleaned as the join key; fall back to raw MPN col.
    join_key_col = header_to_col.get("MPN_cleaned", header_to_col["Manufacture Part Number"])
    meta: dict[str, dict] = {}
    for r in range(2, ws.max_row + 1):
        join_key_raw = ws.cell(r, join_key_col).value
        key = normalize_mpn(join_key_raw)
        if not key or key in meta:
            continue
        meta[key] = {h: ws.cell(r, header_to_col[h]).value for h in CHIP_FIELDS}
    return meta


def read_ok_rows(csv_path: Path) -> list[dict]:
    rows: list[dict] = []
    with csv_path.open(encoding="utf-8-sig", newline="") as f:
        for r in csv.DictReader(f):
            if r.get("status") != "ok":
                continue
            src = r.get("source") or ""
            if any(src.startswith(p) for p in DROP_SOURCE_PREFIXES):
                continue
            r["stockpool_qty"] = parse_qty(r.get("stockpool_qty", ""))
            rows.append(r)
    return rows


def detect_mirror(r: dict) -> bool:
    src = r.get("source") or ""
    wh = (r.get("warehouse") or "")
    if " — mirror" in wh:
        return True
    if src.startswith("FUTURE_") and "(global)" in wh:
        idx = (r.get("warehouse_idx") or "").strip()
        try:
            return int(idx) > 1
        except ValueError:
            return False
    if src.startswith("ELEMENT14_") and wh.startswith("Element14 (") and wh.endswith(")"):
        return True
    return False


def annotate(rows: list[dict], track: str) -> None:
    for r in rows:
        r["track"] = track
        qty = r.get("stockpool_qty")
        r["in_stock"] = bool(qty is not None and qty > 0)
        r["is_mirror"] = detect_mirror(r)


def classify_qty(api_qtys: list, scr_qtys: list) -> str:
    api_set = {q for q in api_qtys if q is not None}
    scr_set = {q for q in scr_qtys if q is not None}
    if not scr_set and not api_set:
        return "both_null"
    if not scr_set or not api_set:
        return "one_side_null"
    if api_set & scr_set:
        return "match"
    return "mismatch"


def build_output_row(raw: dict, chip_meta: dict, columns: list[str]) -> dict:
    """Produce an output dict whose keys are the new column headers.

    Computed columns (Is_orig_manufacture, Is_cheapest, price_rank) are left as
    None here — they're filled in by post-processing passes in main().
    """
    mpn = normalize_mpn(raw.get("input_mpn"))
    meta = chip_meta.get(mpn, {})
    out: dict = {}
    for col in columns:
        # Sheet 3 carries `note`, which is not in COLUMN_SOURCE_MAP.
        if col == "note":
            out[col] = raw.get("note")
            continue
        kind, key = COLUMN_SOURCE_MAP.get(col, (None, None))
        if kind == "merge":
            out[col] = raw.get(key)
        elif kind == "chip":
            value = meta.get(key)
            # v1.10 — `Manufacture Part Number` column shows the raw chip-list
            # MPN. When there's no chip-list match, fall back to input_mpn
            # (the agent-cleaned MPN we searched on) so the cell isn't blank.
            if value is None and col == "Manufacture Part Number":
                value = raw.get("input_mpn")
            out[col] = value
        elif kind == "lead_time":
            out[col] = days_to_weeks(raw.get(key))
        else:  # "computed", "blank", or unknown — fill later or leave None
            out[col] = None
    return out


# ---------- v1.9 computed-column helpers -----------------------------------

_MFR_WAREHOUSE_PARENS = re.compile(r"^Manufacturer warehouse\s*\((.+?)\)\s*$", re.IGNORECASE)


def is_orig_manufacture(warehouse, manufacture) -> bool:
    """True iff the warehouse/vender string represents the manufacturer's own
    supply (vs a third-party distributor).

    Three-step algorithm:
      1. Special-case `"Manufacturer warehouse (X)"` — the parens content IS
         the vendor identity; match Mfr against X (not the prefix).
      2. Otherwise strip any trailing `" (...)"` parenthetical — typically a
         distributor SKU like `DigiKey (497-STM32...)` that would otherwise
         false-positive against Mfr `STM`.
      3. On the cleaned warehouse string:
         - Mfr length ≥ 3: substring match either direction (handles
           "Nexperia 安世" / "Nexperia", "STMicroelectronics" / "STM").
         - Mfr length 1-2 (TI, ON, ST, AD): token-prefix match against runs
           of ≥2 consecutive uppercase letters in the ORIGINAL string.
           This filters out lowercase prepositions ("on" in "Mouser (on
           order)") while preserving real abbreviations ("ON" in
           "ON Semiconductor", "STM" in "STMicroelectronics").
    """
    if not warehouse or not manufacture:
        return False
    wh_str = str(warehouse).strip()
    mfr_str = str(manufacture)
    m = _MFR_WAREHOUSE_PARENS.match(wh_str)
    if m:
        # Match Mfr against the parens content itself.
        return _fuzzy_mfr_vs_wh(m.group(1), mfr_str)
    # General case — strip any trailing distributor-SKU parenthetical.
    wh_clean = re.sub(r"\s*\([^)]*\)\s*$", "", wh_str)
    return _fuzzy_mfr_vs_wh(wh_clean, mfr_str)


def _fuzzy_mfr_vs_wh(warehouse_clean: str, manufacture: str) -> bool:
    wh_norm = re.sub(r"[^A-Z0-9]", "", warehouse_clean.upper())
    mfr_norm = re.sub(r"[^A-Z0-9]", "", str(manufacture).upper())
    if not wh_norm or not mfr_norm:
        return False
    if len(mfr_norm) >= 3:
        return mfr_norm in wh_norm or wh_norm in mfr_norm
    # Short Mfr: only match ≥2-uppercase tokens in the original string.
    wh_upper_tokens = re.findall(r"[A-Z][A-Z0-9]+", warehouse_clean)
    return any(t.startswith(mfr_norm) or mfr_norm.startswith(t) for t in wh_upper_tokens)


def is_cny_currency(value) -> bool:
    """True if the currency string looks like CNY (CNY / RMB / ¥)."""
    if value is None:
        return False
    return str(value).strip().upper() in CNY_LIKE_CURRENCIES


def _strip_vat(value):
    """Divide a numeric (or numeric-string) by VAT_DIVISOR; round to 6 decimals.
    None/empty/non-numeric → returns unchanged. Zero → returned as-is."""
    if value is None or value == "":
        return value
    try:
        num = float(value)
    except (TypeError, ValueError):
        return value
    if num == 0:
        return num
    return round(num / VAT_DIVISOR, 6)


def apply_vat_strip(out_rows: list[dict]) -> int:
    """For CNY/RMB rows, divide the two unit-price columns by 1.13 in place.
    Returns the count of rows touched."""
    n = 0
    for r in out_rows:
        if not is_cny_currency(r.get(TRADE_CURRENCY_HEADER)):
            continue
        before_max = r.get(MAX_PRICE_HEADER)
        before_min = r.get(MIN_PRICE_HEADER)
        r[MAX_PRICE_HEADER] = _strip_vat(before_max)
        r[MIN_PRICE_HEADER] = _strip_vat(before_min)
        if r[MAX_PRICE_HEADER] != before_max or r[MIN_PRICE_HEADER] != before_min:
            n += 1
    return n


def annotate_is_orig_manufacture(out_rows: list[dict]) -> int:
    """Set Is_orig_manufacture per row from Warehouse/vender + Manufacture.
    Returns True-count."""
    n_true = 0
    for r in out_rows:
        flag = is_orig_manufacture(r.get("Warehouse/vender"), r.get("Manufacture"))
        r["Is_orig_manufacture"] = flag
        if flag:
            n_true += 1
    return n_true


def compute_price_ranks(out_rows: list[dict]) -> tuple[int, int]:
    """Group rows by MPN_cleaned_byAgent and:
       - set Is_cheapest=True on the min-price row(s) per MPN (ties → all True)
       - set price_rank using dense rank (1, 2, 2, 3) ascending by price

    Null / 0 prices are EXCLUDED from the comparison: Is_cheapest=False,
    price_rank=None. Operates on Unit price w/o VAT (max qty) — the
    procurement-relevant tier.

    v1.10 — grouping key is `MPN_cleaned_byAgent` (canonical cleaned form)
    rather than `Manufacture Part Number` (raw chip-list MPN), because the
    raw column may vary across rows of the same chip (chip-list match vs
    fallback to input_mpn), which would break per-MPN grouping.

    Returns (n_cheapest_rows, n_mpns_with_ranks).
    """
    by_mpn: dict[str, list[dict]] = defaultdict(list)
    for r in out_rows:
        mpn = r.get("MPN_cleaned_byAgent")
        if mpn:
            by_mpn[mpn].append(r)

    # Initialize defaults
    for r in out_rows:
        r["Is_cheapest"] = False
        r["price_rank"] = None

    n_cheapest = 0
    n_mpns_ranked = 0
    for mpn, rows in by_mpn.items():
        priced = [(r, float(r[MAX_PRICE_HEADER]))
                  for r in rows
                  if _is_positive_number(r.get(MAX_PRICE_HEADER))]
        if not priced:
            continue
        # Dense rank
        unique_sorted_prices = sorted({p for _, p in priced})
        price_to_rank = {p: i + 1 for i, p in enumerate(unique_sorted_prices)}
        min_price = unique_sorted_prices[0]
        for r, p in priced:
            r["price_rank"] = price_to_rank[p]
            if p == min_price:
                r["Is_cheapest"] = True
                n_cheapest += 1
        n_mpns_ranked += 1
    return n_cheapest, n_mpns_ranked


def _is_positive_number(value) -> bool:
    if value is None or value == "":
        return False
    try:
        return float(value) > 0
    except (TypeError, ValueError):
        return False


def is_high_risk_in_stock(out_row: dict) -> bool:
    risk = out_row.get("risk")
    if not isinstance(risk, str):
        if risk is None:
            return False
        risk = str(risk)
    if risk.strip().lower() != "high":
        return False
    qty = out_row.get(QTY_HEADER)
    return isinstance(qty, int) and qty > 0


def _sort_key(out_row: dict) -> tuple:
    """Procurement sort: risk (high first), MPN, Broker, then qty desc.

    v1.10 — sort by `MPN_cleaned_byAgent` (canonical cleaned form) so all
    rows of the same chip cluster together regardless of raw chip-list
    variations. Falls back to `Manufacture Part Number` if cleaned is missing.
    """
    risk = out_row.get("risk")
    risk_str = risk.strip().lower() if isinstance(risk, str) else ""
    rank = RISK_RANK.get(risk_str, 2 if risk_str else 3)
    mpn = out_row.get("MPN_cleaned_byAgent") or out_row.get("Manufacture Part Number") or ""
    broker = out_row.get("Broker name") or ""
    qty = out_row.get(QTY_HEADER)
    qty_neg = -(qty if isinstance(qty, int) else 0)
    return (rank, str(mpn), str(broker), qty_neg)


# ---------- writers ---------------------------------------------------------

def _cell_value(v):
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, (str, int, float)):
        return v
    return str(v)


def _style_workbook(ws, columns: list[str], rows: list[dict]) -> None:
    header_align = Alignment(vertical="center", wrap_text=True)
    for col_idx, col in enumerate(columns, 1):
        c = ws.cell(row=1, column=col_idx, value=col)
        fill, font = _header_style(col_idx, col)
        c.fill = fill
        c.font = font
        c.alignment = header_align
    for row_idx, r in enumerate(rows, 2):
        for col_idx, col in enumerate(columns, 1):
            ws.cell(row=row_idx, column=col_idx, value=_cell_value(r.get(col)))
    ws.freeze_panes = "A2"
    for col_idx, col in enumerate(columns, 1):
        if col in COLUMN_WIDTHS:
            ws.column_dimensions[get_column_letter(col_idx)].width = COLUMN_WIDTHS[col]
            continue
        header_widest = max((len(line) for line in str(col).split("\n")), default=0)
        sample = [str(r.get(col)) for r in rows if r.get(col) is not None][:300]
        sample_widest = max((len(s) for s in sample), default=0)
        width = max(header_widest, sample_widest) + 2
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(width, 10), 50)
    # AutoFilter dropdowns across header + data range.
    last_col = get_column_letter(len(columns))
    last_row = max(2, 1 + len(rows))
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"
    # v1.9 G — hide ref_* + the explicit set on every sheet that carries them.
    _apply_column_hides(ws, columns)


def _apply_row_fills(ws, columns: list[str], rows: list[dict]) -> None:
    ncols = len(columns)
    for row_idx, r in enumerate(rows, 2):
        if r.get("in_stock"):
            fill = GREEN
        elif r.get(QTY_HEADER) == 0:
            fill = GREY
        else:
            continue
        for col_idx in range(1, ncols + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill


def _apply_column_hides(ws, columns: list[str]) -> None:
    """Mark columns hidden per v1.9 G: all ref_* + explicit set."""
    for col_idx, col in enumerate(columns, 1):
        if col.startswith("ref_") or col in HIDDEN_COLUMNS_EXPLICIT:
            ws.column_dimensions[get_column_letter(col_idx)].hidden = True


# ---------- v1.9 I — auxiliary reference sheets ----------------------------

_DATA_DICT_HEADERS = ["Column", "Type", "Description"]
# Descriptions per ref/merged_output_fields_mapping_v3_20260520.xlsx (Description
# column). Blank entries map to blank cells. Order matches OUTPUT_COLUMNS so
# the sheet reads top-to-bottom in the same order as All_data's columns.
_DATA_DICT_ROWS: list[tuple[str, str, str]] = [
    ("Category",                            "text",   "Category"),
    ("Project",                             "text",   "Project name"),
    ("EMS/Finish Goods",                    "text",   "EMS or finished-goods"),
    ("12NC_PCBA",                           "text",   "12NC"),
    ("Manufacture Part Number",             "text",   "Raw MPN as written in the chip list (preserves package suffix / variant descriptors / Chinese annotations)"),
    ("MPN_cleaned_byAgent",                 "text",   "Agent-cleaned MPN actually sent to API/Scraper sources (and used as chip-list join key). See CLAUDE.md Hard Rule #8."),
    ("Manufacture",                         "text",   "Manufacturer Name"),
    ("Quantity",                            "int",    ""),
    ("Currency",                            "text",   ""),
    ("Current Price",                       "number", ""),
    ("Type",                                "text",   "Prescribed or non-Prescribed"),
    ("risk",                                "text",   "Shortage risk tier — 'high' / 'low'"),
    ("in_stock",                            "bool",   "TRUE when this warehouse row reports positive stock"),
    ("Broker name",                         "text",   "Distributor source (e.g., DIGIKEY_得捷电子, LCSC_立创商城, ARROW_艾睿)"),
    ("Data collect method",                 "text",   "Which method is using to collect the info. Scraper refers to web-scraping, API refers to official api"),
    ("Warehouse/vender",                    "text",   "Warehouse name OR vender name"),
    ("Is_orig_manufacture",                 "bool",   "TRUE when Warehouse/vender looks like the manufacturer's own warehouse"),
    ("Is_cheapest",                         "bool",   "TRUE for the row(s) with the minimum Unit price w/o VAT (max qty) for this MPN."),
    ("price_rank",                          "int",    "1..N rank by ascending Unit price w/o VAT (max qty)."),
    ("Stock Location",                      "text",   "Country / region the warehouse ships from"),
    ("Available Quantity",                  "int",    "Units in stock at this warehouse."),
    ("ship infor after order placed",       "text",   "Distributor's SLA / ship-time string"),
    ("Lead Time (Week)",                    "number", "Lead time in weeks"),
    ("MOQ",                                 "int",    "Minimum order quantity"),
    ("Maximum order qty",                   "int",    "Quantity threshold of the top price tier"),
    ("Unit price w/o VAT (max qty)",        "number", "Unit price at the max tier. For CNY/RMB rows, divided by 1.13 to strip Chinese VAT."),
    ("Minimum order qty",                   "int",    "Quantity threshold of the bottom price tier"),
    ("Unit price w/o VAT (min qty)",        "number", "Unit price at the min tier, same VAT treatment as max"),
    ("Number of price tiers",               "int",    "Count of price tiers reported by the source"),
    ("Trade Currency",                      "text",   "Distributor's quoted currency (CNY / USD / EUR / RMB)."),
    ("Date of Code",                        "text",   ""),
    ("Reel/Cut Reel",                       "text",   ""),
    ("Certificate of Conformity(Yes/No)",   "text",   ""),
    ("ref_Warehouse/vender ID",             "int",    ""),
    ("ref_returned_mpn",                    "text",   ""),
    ("ref_vendor_sku",                      "text",   ""),
    ("ref_returned_mfr",                    "text",   ""),
    ("ref_mfr_match",                       "bool",   ""),
    ("ref_is_mirror",                       "bool",   ""),
    ("ref_datasheet_url",                   "text",   ""),
    ("ref_status",                          "text",   ""),
    ("ref_error",                           "text",   ""),
]

_SOURCE_AVAIL_HEADERS = [
    "Source", "Scraper √", "Scraper coverage", "Scraper 可靠性",
    "API √", "API coverage",
]
# v1.9 I — direct port of the TL;DR table from doc/data_sources_overview.md
# minus the "Best use" column. 14 rows. Hardcoded — small data, low churn;
# auto-syncing from the markdown would over-engineer.
_SOURCE_AVAIL_ROWS: list[tuple[str, ...]] = [
    ("DIGIKEY 得捷电子",       "√", "56 %", "较高", "√", "59 %"),
    ("LCSC 立创商城",          "√", "79 %", "中等", "√", "70 %"),
    ("ARROW 艾睿",              "✗", "—",    "—",   "√", "42 %"),
    ("Element14 e络盟",         "✗", "—",    "—",   "√", "43 %"),
    ("Mouser 贸泽",            "✗", "—",    "—",   "√", "58 %"),
    ("买芯片网 (bom2buy.com)", "√", "60 %", "中等", "✗", "—"),
    ("FUTURE 富昌",             "√", "51 %", "较高", "✗", "—"),
    ("HQEW 华强电子网",        "√", "82 %", "较低", "✗", "—"),
    ("ICKEY 云汉芯城",          "√", "80 %", "较低", "✗", "—"),
    ("ONEYAC 唯样商城",        "√", "51 %", "中等", "✗", "—"),
    ("Rochester",               "√", "11 %", "较低", "✗", "—"),
    ("RSONLINE RS 欧时",       "√", "29 %", "较高", "✗", "—"),
    ("Verical (verical.com)",   "✗", "—",    "—",   "✗", "—"),
    ("Chip1Stop (chip1stop.com)", "✗", "—",  "—",   "✗", "—"),
]


def build_data_dictionary_sheet(wb) -> None:
    ws = wb.create_sheet("Data dictionary")
    header_align = Alignment(vertical="center", wrap_text=True)
    body_align = Alignment(vertical="top", wrap_text=True)
    # Header
    for c_idx, h in enumerate(_DATA_DICT_HEADERS, 1):
        c = ws.cell(row=1, column=c_idx, value=h)
        c.fill = HEADER_FILL_BLUE
        c.font = HEADER_FONT_WHITE
        c.alignment = header_align
    # Body
    for r_idx, row in enumerate(_DATA_DICT_ROWS, 2):
        for c_idx, v in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=v)
            cell.alignment = body_align
            cell.font = Font(name="Calibri")
    # Widths
    widths = [32, 10, 75]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def build_source_availability_sheet(wb) -> None:
    ws = wb.create_sheet("Source Availability")
    header_align = Alignment(vertical="center", wrap_text=True, horizontal="center")
    center_align = Alignment(horizontal="center", vertical="center")
    body_align = Alignment(vertical="center")
    for c_idx, h in enumerate(_SOURCE_AVAIL_HEADERS, 1):
        c = ws.cell(row=1, column=c_idx, value=h)
        c.fill = HEADER_FILL_BLUE
        c.font = HEADER_FONT_WHITE
        c.alignment = header_align
    for r_idx, row in enumerate(_SOURCE_AVAIL_ROWS, 2):
        for c_idx, v in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=v)
            cell.font = Font(name="Calibri")
            # First column left-aligned (source name); rest centered.
            cell.alignment = body_align if c_idx == 1 else center_align
    widths = [30, 12, 18, 16, 10, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    # Hide the "Scraper 可靠性" column (col 4) — kept in the sheet for any
    # later business review, but hidden from the default view.
    ws.column_dimensions[get_column_letter(4)].hidden = True
    ws.freeze_panes = "A2"


def write_csv(rows: list[dict], columns: list[str], path: Path) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=columns)
        w.writeheader()
        for r in rows:
            w.writerow({k: ("" if r.get(k) is None else r.get(k)) for k in columns})


# ---------- main ------------------------------------------------------------

def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--env", choices=("test", "prod"), default="test",
                    help="Environment root: 'test' → test/{api,scraper,merged}/ (default), "
                         "'prod' → production/{api,scraper,merged}/.")
    ap.add_argument("--api", type=Path, default=None,
                    help="API BatchTest folder (default: newest under <env_root>/api/)")
    ap.add_argument("--scr", type=Path, default=None,
                    help="Scraper BatchTest folder (default: newest under <env_root>/scraper/)")
    ap.add_argument("--out", type=Path, default=None,
                    help="Output dir (default: <env_root>/merged/Merge_<api_ts>__<scr_ts>/)")
    ap.add_argument("--chip-list", type=Path, default=CHIP_LIST_PATH,
                    help=f"Chip metadata xlsx (default: {CHIP_LIST_PATH})")
    side_group = ap.add_mutually_exclusive_group()
    side_group.add_argument("--api-only", action="store_true",
                            help="Use API batch only; treat scraper side as empty. "
                                 "Cross-validation (Sheet 3) will be empty. "
                                 "Output folder is named Merge_<api_ts>__none/.")
    side_group.add_argument("--scraper-only", action="store_true",
                            help="Use scraper batch only; treat API side as empty. "
                                 "Cross-validation (Sheet 3) will be empty. "
                                 "Output folder is named Merge_none__<scr_ts>/.")
    args = ap.parse_args()

    env_root = ENV_ROOTS[args.env]
    api_root = env_root / "api"
    scr_root = env_root / "scraper"
    merged_root = env_root / "merged"

    # Resolve which sides we read. --api-only / --scraper-only switches off
    # the missing side entirely so the orchestrator can complete a merge with
    # whatever's available after a partial-pipeline failure.
    use_api = not args.scraper_only
    use_scraper = not args.api_only

    api_dir = (args.api or latest_batch(api_root)) if use_api else None
    scr_dir = (args.scr or latest_batch(scr_root)) if use_scraper else None
    api_csv = (api_dir / "batch_index.csv") if api_dir else None
    scr_csv = (scr_dir / "batch_index.csv") if scr_dir else None
    if api_csv is not None and not api_csv.exists():
        raise SystemExit(f"missing {api_csv}")
    if scr_csv is not None and not scr_csv.exists():
        raise SystemExit(f"missing {scr_csv}")

    chip_meta = load_chip_meta(args.chip_list)

    api_ts = re.sub(r"^BatchTest_", "", api_dir.name) if api_dir else "none"
    scr_ts = re.sub(r"^BatchTest_", "", scr_dir.name) if scr_dir else "none"
    out_dir = args.out or (merged_root / f"Merge_{api_ts}__{scr_ts}")
    out_dir.mkdir(parents=True, exist_ok=True)

    api_rows_raw = read_ok_rows(api_csv) if api_csv else []
    scr_rows_raw = read_ok_rows(scr_csv) if scr_csv else []

    # MPN-match filter: drop rows where the channel returned a different MPN
    # than what we searched for (punctuation-insensitive comparison).
    api_rows = [r for r in api_rows_raw if returned_mpn_matches_input(r)]
    scr_rows = [r for r in scr_rows_raw if returned_mpn_matches_input(r)]
    n_api_dropped_mpn = len(api_rows_raw) - len(api_rows)
    n_scr_dropped_mpn = len(scr_rows_raw) - len(scr_rows)

    annotate(api_rows, "api")
    annotate(scr_rows, "scraper")

    # Primary merge — API wins per (mpn, source).
    api_keys = {(r["input_mpn"], r["source"]) for r in api_rows}
    scr_kept = [r for r in scr_rows if (r["input_mpn"], r["source"]) not in api_keys]
    scr_suppressed = [r for r in scr_rows if (r["input_mpn"], r["source"]) in api_keys]
    merged = list(api_rows) + scr_kept

    # Cross-validation — over (mpn, source) where both tracks had data.
    api_by_key: dict[tuple, list[dict]] = defaultdict(list)
    for r in api_rows:
        api_by_key[(r["input_mpn"], r["source"])].append(r)
    scr_by_key: dict[tuple, list[dict]] = defaultdict(list)
    for r in scr_suppressed:
        scr_by_key[(r["input_mpn"], r["source"])].append(r)

    mismatch_rows: list[dict] = []
    for key, scr_group in scr_by_key.items():
        api_group = api_by_key[key]
        verdict = classify_qty(
            [r["stockpool_qty"] for r in api_group],
            [r["stockpool_qty"] for r in scr_group],
        )
        if verdict != "mismatch":
            continue
        api_max = max((q for q in (r["stockpool_qty"] for r in api_group) if q is not None), default=0)
        scr_max = max((q for q in (r["stockpool_qty"] for r in scr_group) if q is not None), default=0)
        note = f"API max qty {api_max} vs scraper max {scr_max}"
        for r in scr_group:
            r2 = dict(r)
            r2["note"] = note
            mismatch_rows.append(r2)

    # Project RAW rows → OUTPUT rows (new column names, lead-time converted,
    # chip-list enrichment applied).
    sheet2_rows = [build_output_row(r, chip_meta, OUTPUT_COLUMNS) for r in merged]
    sheet3_rows = [build_output_row(r, chip_meta, MISMATCH_COLUMNS) for r in mismatch_rows]

    # v1.9 post-processing — ORDER MATTERS:
    #   1. apply VAT-strip FIRST (computes are pre-VAT semantically: cheapest
    #      comparison should be in stripped numbers).
    #   2. annotate Is_orig_manufacture (row-local — order irrelevant but
    #      grouped for clarity).
    #   3. compute Is_cheapest / price_rank from the now-stripped prices.
    n_vat_stripped = apply_vat_strip(sheet2_rows)
    apply_vat_strip(sheet3_rows)  # Sheet 3 inherits VAT-strip (no return needed)
    n_is_orig_true = annotate_is_orig_manufacture(sheet2_rows)
    n_cheapest_rows, n_mpns_ranked = compute_price_ranks(sheet2_rows)

    # Unified sort across all three sheets: risk → MPN → Broker → qty desc.
    sheet2_rows.sort(key=_sort_key)
    sheet3_rows.sort(key=_sort_key)

    # Sheet 1 is a strict subset of sheet 2: risk=high AND qty>0.
    sheet1_rows = [r for r in sheet2_rows if is_high_risk_in_stock(r)]
    sheet1_rows.sort(key=_sort_key)

    # Write xlsx.
    today = date.today().strftime("%Y%m%d")
    out_stem = f"Versuni_chip_stock_availability_check_{today}"
    xlsx_path = out_dir / f"{out_stem}.xlsx"
    wb = openpyxl.Workbook()

    # Sheet order: High_risk_positive_stock → All_data → ref_scraper_api_diff
    # → Data dictionary → Source Availability.
    # Sheets 1 and 3 are hidden by default (v1.9 H); Sheet 2 is the default active.
    ws1 = wb.active
    ws1.title = "High_risk_positive_stock"
    _style_workbook(ws1, OUTPUT_COLUMNS, sheet1_rows)
    # Sheet 1 = all in-stock by definition → no per-row green fill (uninformative).

    ws2 = wb.create_sheet("All_data")
    _style_workbook(ws2, OUTPUT_COLUMNS, sheet2_rows)
    _apply_row_fills(ws2, OUTPUT_COLUMNS, sheet2_rows)

    ws3 = wb.create_sheet("ref_scraper_api_diff")
    _style_workbook(ws3, MISMATCH_COLUMNS, sheet3_rows)

    # v1.9 I — reference sheets.
    build_data_dictionary_sheet(wb)
    build_source_availability_sheet(wb)

    # v1.9 H — hide Sheet 1 + Sheet 3; make Sheet 2 the default active.
    for sn in HIDDEN_SHEETS:
        if sn in wb.sheetnames:
            wb[sn].sheet_state = "hidden"
    wb.active = wb.sheetnames.index("All_data")

    wb.save(xlsx_path)

    # Companion CSV (= sheet 2).
    csv_path = out_dir / f"{out_stem}.csv"
    write_csv(sheet2_rows, OUTPUT_COLUMNS, csv_path)

    # Run summary.
    n_api_ok = len(api_rows)
    n_scr_ok_total = len(scr_rows)
    n_scr_kept = len(scr_kept)
    n_scr_suppressed = len(scr_suppressed)
    # v1.10 — count chips by MPN_cleaned_byAgent (canonical cleaned form), since
    # `Manufacture Part Number` is now the raw chip-list MPN which may differ
    # between chip-list-matched rows and no-chip-list-match fallback rows for
    # the same chip.
    chips_high_risk = len({r.get("MPN_cleaned_byAgent") for r in sheet1_rows})
    chips_total_merged = len({r.get("MPN_cleaned_byAgent") for r in sheet2_rows})
    chips_matched = sum(1 for m in {r.get("MPN_cleaned_byAgent") for r in sheet2_rows}
                        if m and normalize_mpn(m) in chip_meta)
    print(f"Chip list (Part List Modify)        : {len(chip_meta)} unique MPNs")
    if args.scraper_only:
        print(f"API rows                            : SKIPPED (--scraper-only)")
    else:
        print(f"API rows (status=ok, !HQEW)         : {n_api_ok}  (dropped {n_api_dropped_mpn} for returned_mpn mismatch)")
    if args.api_only:
        print(f"Scraper rows                        : SKIPPED (--api-only)")
    else:
        print(f"Scraper rows (status=ok, !HQEW)    : {n_scr_ok_total}  (dropped {n_scr_dropped_mpn} for returned_mpn mismatch)")
    print(f"  - kept (API has no coverage)     : {n_scr_kept}")
    print(f"  - suppressed (API took over)     : {n_scr_suppressed}")
    print(f"Merged total (Sheet 2 'All_data')   : {len(sheet2_rows)} ({chips_total_merged} MPNs)")
    print(f"  - matched against chip list      : {chips_matched}")
    print(f"  - no chip-list row               : {chips_total_merged - chips_matched}")
    print(f"Sheet 1 'High_risk_positive_stock' : {len(sheet1_rows)} rows, {chips_high_risk} MPNs")
    print(f"Sheet 3 'ref_scraper_api_diff'      : {len(sheet3_rows)} rows")
    print(f"v1.9 — VAT-stripped (CNY/RMB)       : {n_vat_stripped} rows")
    print(f"v1.9 — Is_orig_manufacture=True     : {n_is_orig_true} / {len(sheet2_rows)} rows")
    print(f"v1.9 — Is_cheapest=True             : {n_cheapest_rows} rows over {n_mpns_ranked} MPNs with priced rows")
    print()
    print(f"Written:")
    print(f"  {xlsx_path}")
    print(f"  {csv_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
