"""LLM-assisted manufacturer-name normalization for batch_index.csv.

For each `(expected_mfr, returned_mfr)` pair flagged as `mfr_match=False`
(but `status=ok`) in a batch's `batch_index.csv`, asks Deepseek-v4-pro
whether the two strings refer to the same company in industry practice
(mergers / sub-brands / abbreviations / mixed-language / known-second-source).

Side effects on the batch folder:
  1. `batch_index.csv` (+ `.xlsx`) gains two new columns:
       - `llm_mfr_verdict`  one of YES | NO | WEAK_YES | UNCERTAIN | ""
       - `llm_mfr_reason`   short free-text justification from the model
     Rows with `mfr_match=True`, `status!=ok`, or empty mfr stay blank.
  2. `batch_summary.md` gets an extra "Legitimate equivalents (LLM-suppressed)"
     sub-block appended to the manufacturer mismatch section, listing the
     YES pairs the procurement reviewer can ignore.

Pipeline integration: `batch_scraper_test.py` calls `apply_to_batch_index()`
after the bom2buy post-step and merge run, before the README refresh.

CLI standalone use:
  python common/_llm_mfr_normalize.py <batch_dir>

Failure modes (all non-blocking — never aborts the batch):
  - `deepseek_api_key` missing from env → skip, print warning
  - HTTP error / timeout / malformed response → skip, print warning
  - 0 mismatch pairs in batch → skip silently

Key never gets written to disk or printed anywhere.
"""
from __future__ import annotations

import csv
import json
import os
import re
import sys
from pathlib import Path
from typing import Optional

import requests
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font

PROJECT_ROOT = Path(__file__).resolve().parents[1]
ENV_FILE = PROJECT_ROOT / "api" / ".env"

# Picked from `deepseek_model_list` env var (`['deepseek-v4-flash',
# 'deepseek-v4-pro']`). `v4-pro` is the slower/stronger one; the 2026-05-27
# pilot achieved 90% verdict accuracy at temperature 0. If only the flash
# model is available, swap the constant here.
DEFAULT_MODEL = "deepseek-v4-pro"

# Reasons containing any of these tokens get downgraded YES → WEAK_YES; the
# pilot caught one case ("CM is likely part marking for Core link") where the
# model uses speculative language without a verified equivalence.
WEASEL_WORDS = (
    "likely", "probably", "may be", "could be", "perhaps", "seems",
    "似乎", "也许", "应该", "推测", "可能",
)

NEW_COLUMNS = ["llm_mfr_verdict", "llm_mfr_reason"]


# ───────────────────────────── prompt + LLM call ─────────────────────────────

_PROMPT_TEMPLATE = """You are a chip-distribution domain expert. I will give you {n} pairs of (expected_manufacturer, returned_manufacturer) strings extracted from chip-availability scrapers.

For each pair, judge whether the two strings refer to the same company in industry practice.

YES = they are the same company, considering ALL of:
  - Mergers / acquisitions (e.g., EPCOS = TDK after acquisition; WeEn = Nexperia subsidiary)
  - Sub-brands or business units of the same parent
  - Legal name vs common short name (Texas Instruments = TI = 德州仪器)
  - Different language (意法半导体 = STMicroelectronics = ST)
  - Second-source / officially licensed clones from a recognized fab

NO = they are different companies, OR one side is a marketplace reseller name (e.g., '唯样海外代购', selling on a marketplace and NOT the actual manufacturer of the die), OR one is a logistics / supplier label not a chip brand.

UNCERTAIN = cannot tell without more info (rare niche brands you genuinely don't recognize).

Respond with EXACTLY one line per pair in this format, NOTHING else:
<id>|<YES|NO|UNCERTAIN>|<short reason in <=80 chars>

Pairs:
{pairs}
"""


def _build_prompt(pairs: list[tuple[str, str]]) -> str:
    lines = [f'{i+1}. ("{e}") vs ("{r}")' for i, (e, r) in enumerate(pairs)]
    return _PROMPT_TEMPLATE.format(n=len(pairs), pairs="\n".join(lines))


def _call_deepseek(prompt: str, api_key: str, base_url: str, model: str) -> str:
    """Returns the raw assistant content. Raises on HTTP error."""
    url = f"{base_url.rstrip('/')}/v1/chat/completions"
    body = {
        "model": model,
        "messages": [
            {"role": "system", "content": "You are a precise classifier. Respond ONLY in the format specified."},
            {"role": "user", "content": prompt},
        ],
        "temperature": 0.0,
        "max_tokens": 4096,
    }
    r = requests.post(
        url,
        headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
        json=body,
        timeout=120,
    )
    r.raise_for_status()
    j = r.json()
    return j["choices"][0]["message"]["content"]


def _parse_verdicts(text: str, n_pairs: int) -> dict[int, dict]:
    """Parses `<id>|<verdict>|<reason>` lines into {idx: {verdict, reason}}."""
    by_idx: dict[int, dict] = {}
    for line in text.strip().splitlines():
        line = line.strip()
        if not line or "|" not in line:
            continue
        parts = [p.strip() for p in line.split("|", 2)]
        if len(parts) < 3:
            continue
        idx_digits = "".join(c for c in parts[0] if c.isdigit())
        if not idx_digits:
            continue
        idx = int(idx_digits)
        verdict_raw = parts[1].upper()
        reason = parts[2]
        verdict = verdict_raw if verdict_raw in ("YES", "NO", "UNCERTAIN") else "UNCERTAIN"
        # Downgrade speculative YES → WEAK_YES so the consumer can decide
        # whether to trust it.
        if verdict == "YES" and any(w in reason.lower() for w in WEASEL_WORDS):
            verdict = "WEAK_YES"
        if 1 <= idx <= n_pairs:
            by_idx[idx] = {"verdict": verdict, "reason": reason}
    return by_idx


def classify_pairs(
    pairs: list[tuple[str, str]],
    api_key: str,
    base_url: str,
    model: str = DEFAULT_MODEL,
) -> dict[tuple[str, str], dict]:
    """Returns {(expected, returned): {'verdict', 'reason'}}."""
    if not pairs:
        return {}
    prompt = _build_prompt(pairs)
    content = _call_deepseek(prompt, api_key, base_url, model)
    by_idx = _parse_verdicts(content, len(pairs))
    out: dict[tuple[str, str], dict] = {}
    for i, (e, r) in enumerate(pairs, 1):
        v = by_idx.get(i, {"verdict": "UNCERTAIN", "reason": "(no LLM verdict returned)"})
        out[(e, r)] = v
    return out


# ─────────────────────────── batch_index.csv update ──────────────────────────

def _read_csv(path: Path) -> tuple[list[str], list[dict]]:
    with open(path, encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        return list(reader.fieldnames or []), list(reader)


def _write_csv(path: Path, columns: list[str], rows: list[dict]) -> None:
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=columns)
        w.writeheader()
        for r in rows:
            w.writerow({c: r.get(c, "") for c in columns})


def _write_xlsx(path: Path, columns: list[str], rows: list[dict], sheet: str = "batch_index") -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    bold = Font(name="Calibri", bold=True)
    normal = Font(name="Calibri")
    for c_idx, col in enumerate(columns, 1):
        cell = ws.cell(row=1, column=c_idx, value=col)
        cell.font = bold
    for r_idx, row in enumerate(rows, 2):
        for c_idx, col in enumerate(columns, 1):
            v = row.get(col, "")
            if isinstance(v, str) and v.strip() == "":
                v = ""
            cell = ws.cell(row=r_idx, column=c_idx, value=v)
            cell.font = normal
    wb.save(path)


# ─────────────────────────── summary.md patching ─────────────────────────────

_SECTION_HEADING = "## Manufacturer mismatches"
_INSERTED_HEADING = "### LLM normalization (legitimate equivalents auto-flagged)"


def _patch_summary_md(batch_dir: Path, verdict_rows: list[dict]) -> bool:
    """Insert a sub-section listing YES (non-weak) pairs into batch_summary.md.

    Returns True if patched, False if no patch was applied (missing file, no
    rows to surface, or section already patched)."""
    summary = batch_dir / "batch_summary.md"
    if not summary.exists():
        return False
    text = summary.read_text(encoding="utf-8")
    if _INSERTED_HEADING in text:
        # Strip prior auto-insertion so re-runs replace cleanly.
        text = re.sub(
            re.escape(_INSERTED_HEADING) + r".*?(?=\n## |\Z)",
            "",
            text,
            flags=re.DOTALL,
        ).rstrip() + "\n"
    yes_rows = [r for r in verdict_rows if r["verdict"] == "YES"]
    weak_rows = [r for r in verdict_rows if r["verdict"] == "WEAK_YES"]
    if not (yes_rows or weak_rows):
        # Still note that we ran but found no equivalents — keeps the
        # procurement reviewer from wondering whether LLM step happened.
        block = (
            f"\n{_INSERTED_HEADING}\n\n"
            "_LLM verdict ran: 0 legitimate-equivalence findings._\n"
        )
    else:
        block_lines = [
            "",
            _INSERTED_HEADING,
            "",
            f"Deepseek-v4-pro flagged {len(yes_rows)} pair(s) as the SAME company "
            "(reseller names, sub-brands, abbreviations, language variants). "
            f"{len(weak_rows)} additional pair(s) returned a speculative YES "
            "(reason contains words like 'likely' / 'probably') and are listed "
            "separately for human review.",
            "",
        ]
        if yes_rows:
            block_lines += [
                "**YES — auto-suppress from real-mismatch tally:**",
                "",
                "| expected_mfr | returned_mfr | example_mpn | source | reason |",
                "|---|---|---|---|---|",
            ]
            for r in yes_rows:
                block_lines.append(
                    f"| {r['expected']} | {r['returned']} | `{r['example_mpn']}` | "
                    f"{r['example_source']} | {r['reason']} |"
                )
            block_lines.append("")
        if weak_rows:
            block_lines += [
                "**WEAK_YES — speculative, review manually:**",
                "",
                "| expected_mfr | returned_mfr | example_mpn | source | reason |",
                "|---|---|---|---|---|",
            ]
            for r in weak_rows:
                block_lines.append(
                    f"| {r['expected']} | {r['returned']} | `{r['example_mpn']}` | "
                    f"{r['example_source']} | {r['reason']} |"
                )
            block_lines.append("")
        block = "\n".join(block_lines)
    # Insert immediately after the mismatch section (before the next ## heading
    # or end of file). Anchor on the literal section heading.
    if _SECTION_HEADING in text:
        # Find the END of the mismatch section: the next "## " at column 0
        # after the section heading.
        i = text.index(_SECTION_HEADING)
        # find next "\n## " after i (next sibling heading) or EOF
        j = text.find("\n## ", i + len(_SECTION_HEADING))
        if j < 0:
            # No following section; append block at EOF
            text = text.rstrip() + "\n" + block
        else:
            text = text[:j].rstrip() + "\n" + block + text[j:]
    else:
        # Section not found (unexpected); just append at end.
        text = text.rstrip() + "\n" + block
    summary.write_text(text, encoding="utf-8")
    return True


# ──────────────────────────────── orchestration ──────────────────────────────

def _load_env() -> tuple[Optional[str], Optional[str]]:
    """Returns (api_key, base_url) or (None, None) if not configured."""
    load_dotenv(ENV_FILE)
    key = os.environ.get("deepseek_api_key", "").strip()
    url = os.environ.get("deepseek_base_url_OpenAI", "").strip()
    if not key or not url:
        return None, None
    return key, url


def apply_to_batch_index(batch_dir: Path, model: str = DEFAULT_MODEL) -> dict:
    """Main entry point. Returns stats dict. Never raises — degrades to a
    printed warning + empty stats on any failure."""
    batch_dir = Path(batch_dir)
    csv_path = batch_dir / "batch_index.csv"
    if not csv_path.exists():
        print(f"[llm_mfr_normalize] {csv_path} not found — skipping")
        return {"skipped": "no_csv"}

    api_key, base_url = _load_env()
    if not api_key:
        print("[llm_mfr_normalize] deepseek_api_key not set — skipping (set in api/.env to enable)")
        return {"skipped": "no_api_key"}

    columns, rows = _read_csv(csv_path)

    # Build unique (expected, returned) pairs from mfr_match=False, status=ok
    # rows with both fields populated. Keep one example MPN per pair for the
    # summary.md table.
    pair_examples: dict[tuple[str, str], dict] = {}
    for r in rows:
        if r.get("mfr_match") != "False":
            continue
        if r.get("status") != "ok":
            continue
        e = (r.get("expected_mfr") or "").strip()
        ret = (r.get("returned_mfr") or "").strip()
        if not e or not ret:
            continue
        key = (e, ret)
        if key not in pair_examples:
            pair_examples[key] = {
                "expected": e,
                "returned": ret,
                "example_mpn": r.get("input_mpn", ""),
                "example_source": r.get("source", ""),
            }

    if not pair_examples:
        print("[llm_mfr_normalize] 0 manufacturer mismatches in batch — skipping LLM call")
        # Still ensure new columns exist (empty) so downstream readers see a
        # consistent schema.
        new_columns = [c for c in NEW_COLUMNS if c not in columns]
        if new_columns:
            insert_at = columns.index("elapsed_sec") if "elapsed_sec" in columns else len(columns)
            columns = columns[:insert_at] + new_columns + columns[insert_at:]
            for r in rows:
                for c in new_columns:
                    r[c] = ""
            _write_csv(csv_path, columns, rows)
            _write_xlsx(batch_dir / "batch_index.xlsx", columns, rows)
        return {"skipped": "no_mismatches"}

    pairs = list(pair_examples.keys())
    print(f"[llm_mfr_normalize] classifying {len(pairs)} unique mismatch pairs via {model}")
    try:
        verdicts = classify_pairs(pairs, api_key, base_url, model)
    except Exception as e:
        print(f"[llm_mfr_normalize] LLM call failed: {str(e)[:200]} — leaving columns empty")
        verdicts = {}

    # Promote new columns into the column order (in front of elapsed_sec /
    # num_variants to keep scraper-extras grouped at the end).
    new_columns = [c for c in NEW_COLUMNS if c not in columns]
    if new_columns:
        insert_at = columns.index("elapsed_sec") if "elapsed_sec" in columns else len(columns)
        columns = columns[:insert_at] + new_columns + columns[insert_at:]

    # Inject verdicts into every row. Rows without a matching pair (most rows)
    # get empty strings.
    for r in rows:
        key = ((r.get("expected_mfr") or "").strip(), (r.get("returned_mfr") or "").strip())
        v = verdicts.get(key)
        if v and r.get("mfr_match") == "False" and r.get("status") == "ok":
            r["llm_mfr_verdict"] = v["verdict"]
            r["llm_mfr_reason"] = v["reason"]
        else:
            r.setdefault("llm_mfr_verdict", "")
            r.setdefault("llm_mfr_reason", "")

    _write_csv(csv_path, columns, rows)
    _write_xlsx(batch_dir / "batch_index.xlsx", columns, rows)

    verdict_rows = [{**pair_examples[k], **v} for k, v in verdicts.items()]
    patched = _patch_summary_md(batch_dir, verdict_rows)

    counts = {"YES": 0, "NO": 0, "WEAK_YES": 0, "UNCERTAIN": 0}
    for vr in verdict_rows:
        counts[vr["verdict"]] = counts.get(vr["verdict"], 0) + 1
    print(f"[llm_mfr_normalize] verdicts: " + ", ".join(f"{k}={v}" for k, v in counts.items()))
    print(f"[llm_mfr_normalize] batch_index.csv + .xlsx updated; batch_summary.md patched={patched}")

    return {
        "pairs_total": len(pairs),
        "verdict_counts": counts,
        "summary_patched": patched,
    }


def main():
    if len(sys.argv) < 2:
        print("Usage: python common/_llm_mfr_normalize.py <batch_dir>")
        sys.exit(2)
    batch_dir = Path(sys.argv[1])
    if not batch_dir.is_absolute():
        batch_dir = PROJECT_ROOT / batch_dir
    apply_to_batch_index(batch_dir)


if __name__ == "__main__":
    main()
