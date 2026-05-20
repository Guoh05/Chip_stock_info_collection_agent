"""Batch-run every chip in the master sheet through the 9 working scrapers
(LCSC, Digikey, HQEW, Future, RSONLINE, ONEYAC, ICKEY, Rochester + bom2buy
as a post-step) and consolidate results.

The first 8 channels run in parallel per chip via a ThreadPoolExecutor, each
as a SUBPROCESS so a hung browser or crashed scraper is killed cleanly with a
hard wallclock timeout. The 9th channel (bom2buy) runs SEQUENTIALLY as a
post-step because it drives the user's Opera browser via Playwright with
exclusive user-data-dir lock — it cannot run alongside other channels.

Per-channel timeouts (seconds, observed-budget × 1.4 safety):
    LCSC v3     240
    Digikey     180
    HQEW         90
    Future      300
    RSONLINE     90
    ONEYAC      120
    ICKEY       150
    Rochester   180
    bom2buy     post-step, in-script subprocess (1 h hard cap)

Master input (2026-05-20+):
    ref/Shortage Emergency Response List_v2.xlsx, sheet `Part List Modify`,
    column `Manufacture Part Number`. Mfr column `Manufacture` is reference
    only. MPNs are deduped (107 unique from 280 raw rows in v2).
    Legacy `Chip_DataSource_Master.xlsx` (header on row 4) is still supported
    if explicitly passed via --xlsx — `load_chip_list` detects and falls back.

Outputs under `test/scraper_test/BatchTest_<YYYYMMDD>_<HH_MM_SS>/`:
    batch_summary.md                — TL;DR + per-channel stats + highlights
    batch_index.csv / .xlsx         — v3 long form (warehouse-exploded)
    batch_index.json                — machine-readable long form
    batch_input.csv                 — verbatim (MPN, expected_mfr) from xlsx
    failures.md                     — non-ok rows grouped by channel
    Test_<sanitized_mpn>_<CHANNEL>/ — per-MPN-per-channel run folder
                                       (populated by the scraper subprocess)

Usage:
    .venv/Scripts/python.exe scraper/scripts/batch_scraper_test.py
    .venv/Scripts/python.exe scraper/scripts/batch_scraper_test.py --limit 3
    .venv/Scripts/python.exe scraper/scripts/batch_scraper_test.py --only LCSC,HQEW
    .venv/Scripts/python.exe scraper/scripts/batch_scraper_test.py --resume
    .venv/Scripts/python.exe scraper/scripts/batch_scraper_test.py --no-bom2buy

bom2buy preconditions (default ON, --no-bom2buy to skip):
    1. Open Opera, navigate to https://www.bom2buy.com/, solve IconCaptcha once.
    2. FULLY close Opera (Task Manager → kill all opera.exe if needed).
    3. Run this script.
    If the captcha session is expired at run time the batch still completes;
    bom2buy can be backfilled later with scrape_bom2buy.py + _merge_bom2buy_into_batch.py.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import subprocess
import sys
import time
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parents[2]
SCRIPTS_DIR = PROJECT_ROOT / "scraper" / "scripts"
SCRAPER_TEST_ROOT = PROJECT_ROOT / "test" / "scraper_test"
DEFAULT_XLSX = PROJECT_ROOT / "ref" / "Shortage Emergency Response List_v2.xlsx"
DEFAULT_SHEET = "Part List Modify"
DEFAULT_MPN_COL = "Manufacture Part Number"
DEFAULT_MFR_COL = "Manufacture"

CHANNELS: dict[str, dict[str, Any]] = {
    "LCSC":      {"script": "scrape_lcsc_v3.py",  "timeout": 240},
    "DIGIKEY":   {"script": "scrape_digikey.py",  "timeout": 180},
    "HQEW":      {"script": "scrape_hqew.py",     "timeout": 90},
    "FUTURE":    {"script": "scrape_future.py",   "timeout": 300},
    "RSONLINE":  {"script": "scrape_rsonline.py", "timeout": 90},
    "ONEYAC":    {"script": "scrape_oneyac.py",   "timeout": 120},
    "ICKEY":     {"script": "scrape_ickey.py",    "timeout": 150},
    "ROCHESTER": {"script": "scrape_rochester.py","timeout": 180},
}

THROTTLE_SECONDS = 1.0


# --- helpers ---------------------------------------------------------------


def _safe_folder(mpn: str) -> str:
    return re.sub(r"[^A-Za-z0-9._-]", "_", mpn) or "UNKNOWN"


def _looks_like_real_mpn(value: Any) -> bool:
    if value is None:
        return False
    s = str(value).strip()
    if len(s) < 3:
        return False
    if "缺失" in s or "TBD" in s.upper() or "TODO" in s.upper():
        return False
    if not re.search(r"[A-Za-z0-9]", s):
        return False
    return True


def _norm_mfr(s: str | None) -> str:
    if not s:
        return ""
    return re.sub(r"[^A-Z0-9]", "", s.upper())


def _mfr_match(expected: str | None, returned: str | None) -> bool:
    e = _norm_mfr(expected)
    r = _norm_mfr(returned)
    if not e or not r:
        return False
    return e in r or r in e


def _parse_price_str(price_str) -> float | None:
    if price_str is None or price_str == "":
        return None
    if isinstance(price_str, (int, float)):
        return float(price_str)
    s = re.sub(r"[^\d.,\-]", "", str(price_str).strip())
    if not s:
        return None
    if "," in s and "." not in s:
        s = s.replace(",", ".")
    else:
        s = s.replace(",", "")
    try:
        return float(s)
    except ValueError:
        return None


# --- input loading ---------------------------------------------------------


def _clean_mpn_str(s: object) -> str:
    """Strip whitespace including non-breaking spaces (U+00A0) that sneak into
    MPNs copy-pasted from Excel."""
    if s is None:
        return ""
    return str(s).strip().strip("\xa0").strip()


def load_chip_list(
    xlsx_path: Path,
    sheet: str = DEFAULT_SHEET,
    mpn_col: str = DEFAULT_MPN_COL,
    mfr_col: str = DEFAULT_MFR_COL,
) -> tuple[list[dict], list[dict]]:
    """Read chip list from a header-row-1 sheet, dedup by MPN.

    The 2026-05-20+ master is `ref/Shortage Emergency Response List_v2.xlsx`,
    sheet `Part List Modify` — header on row 1, data from row 2. MPN col is
    `Manufacture Part Number`, mfr col is `Manufacture`. The mfr value is
    reference-only (we keep the first occurrence's mfr after dedup).

    Returns (chips, skipped). `chips` is one entry per UNIQUE MPN; `skipped`
    captures rows whose MPN looks like a placeholder, plus duplicate-MPN rows
    (with reason="duplicate; first kept" — useful for audit).

    Backward-compat: the legacy `ref/Chip_DataSource_Master.xlsx` had header on
    row 4 (so data from row 5) with MPN in col 1 and mfr in col 2. Pass
    `sheet=None` to fall back to that layout, OR just keep using the legacy
    file and override `--xlsx`.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    if sheet is None or sheet not in wb.sheetnames:
        # Legacy layout: first sheet, header row 4, MPN col 1, mfr col 2.
        ws = wb.worksheets[0]
        rows_iter = ((r, ws.cell(row=r, column=1).value, ws.cell(row=r, column=2).value)
                     for r in range(5, ws.max_row + 1))
    else:
        ws = wb[sheet]
        # Header row 1: build a column-name → index map
        header = {str(c.value).strip(): c.column_letter for c in ws[1] if c.value}
        if mpn_col not in header:
            raise KeyError(f"Column {mpn_col!r} not found in sheet {sheet!r}. "
                           f"Available: {list(header.keys())}")
        mpn_letter = header[mpn_col]
        mfr_letter = header.get(mfr_col)
        rows_iter = (
            (r,
             ws[f"{mpn_letter}{r}"].value,
             ws[f"{mfr_letter}{r}"].value if mfr_letter else None)
            for r in range(2, ws.max_row + 1)
        )

    chips: list[dict] = []
    skipped: list[dict] = []
    seen_mpns: set[str] = set()
    for r, mpn, mfr in rows_iter:
        mpn_s = _clean_mpn_str(mpn)
        mfr_s = _clean_mpn_str(mfr)
        if not _looks_like_real_mpn(mpn_s):
            if mpn_s or mfr_s:  # skip silently-empty rows; capture non-empty placeholders
                skipped.append({"row": r, "raw_mpn": mpn_s, "raw_mfr": mfr_s,
                                "reason": "missing or non-MPN placeholder"})
            continue
        if mpn_s in seen_mpns:
            skipped.append({"row": r, "raw_mpn": mpn_s, "raw_mfr": mfr_s,
                            "reason": "duplicate; first occurrence kept"})
            continue
        seen_mpns.add(mpn_s)
        chips.append({"row": r, "input_mpn": mpn_s, "expected_mfr": mfr_s})
    return chips, skipped


# --- subprocess execution --------------------------------------------------


def run_subprocess(channel: str, mpn: str, out_dir: Path) -> dict:
    """Run a single scraper subprocess with hard wallclock timeout.

    Returns {"status": ok|exit_N|timeout|exception, "elapsed_sec", "stdout_tail",
    "stderr_tail", "error"}.
    """
    cfg = CHANNELS[channel]
    script = SCRIPTS_DIR / cfg["script"]
    timeout = cfg["timeout"]
    cmd = [sys.executable, str(script), mpn, str(out_dir)]
    # Force UTF-8 in the subprocess so prints of '¥' / Chinese characters don't
    # crash under Windows GBK when stdout is being captured by us. The scrapers'
    # actual data is written to disk independently — only their console prints
    # are affected — but a Unicode crash in a final print still aborts the
    # process with exit code 1, which we'd misclassify as a failure.
    env = {**os.environ, "PYTHONIOENCODING": "utf-8", "PYTHONUTF8": "1"}
    t0 = time.time()
    try:
        proc = subprocess.run(
            cmd,
            timeout=timeout,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            env=env,
        )
        elapsed = time.time() - t0
        if proc.returncode == 0:
            status = "ok"
            error = None
        else:
            status = f"exit_{proc.returncode}"
            error = (proc.stderr or "").strip().splitlines()[-1] if proc.stderr else None
        return {
            "status": status,
            "elapsed_sec": round(elapsed, 1),
            "stdout_tail": (proc.stdout or "")[-2000:],
            "stderr_tail": (proc.stderr or "")[-2000:],
            "error": error,
        }
    except subprocess.TimeoutExpired as exc:
        return {
            "status": "timeout",
            "elapsed_sec": round(time.time() - t0, 1),
            "stdout_tail": (exc.stdout or "")[-2000:] if isinstance(exc.stdout, str) else "",
            "stderr_tail": (exc.stderr or "")[-2000:] if isinstance(exc.stderr, str) else "",
            "error": f"timeout after {timeout}s",
        }
    except Exception as exc:
        return {
            "status": "exception",
            "elapsed_sec": round(time.time() - t0, 1),
            "stdout_tail": "",
            "stderr_tail": "",
            "error": f"{type(exc).__name__}: {exc}",
        }


# --- post-subprocess: load + pick best variant ----------------------------


def load_record(out_dir: Path, mpn: str) -> dict | None:
    """Find the parent JSON the scraper wrote. Scrapers name it `<safe>.json`."""
    safe = _safe_folder(mpn)
    candidate = out_dir / f"{safe}.json"
    if candidate.exists():
        try:
            return json.loads(candidate.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            return None
    # Fallback: any top-level .json
    for p in sorted(out_dir.glob("*.json")):
        if p.name.startswith("_"):
            continue
        try:
            return json.loads(p.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            continue
    return None


def pick_best_extracted(record: dict, channel: str, out_dir: Path, input_mpn: str) -> tuple[dict | None, int]:
    """Choose the variant whose data best represents this MPN.

    Returns (extracted_dict, num_variants). For single-result channels
    (Digikey, HQEW) num_variants is 1 when ok else 0.
    """
    if not record:
        return None, 0

    # Single-variant channels — record carries a flat `extracted` directly.
    # (HQEW, Digikey, and the 4 newer marketplace/aggregator scrapers built
    # in 2026-05-18: RSONLINE, ONEYAC, ICKEY, ROCHESTER.)
    if channel in ("DIGIKEY", "HQEW", "RSONLINE", "ONEYAC", "ICKEY", "ROCHESTER"):
        ex = record.get("extracted")
        if ex:
            # HQEW's extracted may carry an inner variants list — count is informative
            n = len(ex.get("variants") or []) if channel == "HQEW" else 1
            return ex, n if n else 1
        return None, 0

    # LCSC v3 / Future — record["variants"] list
    variants = record.get("variants") or []
    if not variants:
        return None, 0

    target = (input_mpn or "").strip().lower()
    # Alphanumeric-normalized form of the input MPN — used for fuzzy
    # substring containment when the source returned no exact-MPN match.
    # Without this guard, multi-variant sources (LCSC, Future) sometimes
    # return dozens of keyword-tagged unrelated parts and the stock-tiebreak
    # silently picks the highest-stock one (e.g. LTW-M140SXT57-PA → a JST
    # connector with 107k stock). Per "uncertain → blank, never invent",
    # require alphanumeric-substring containment in BOTH directions before
    # falling back; otherwise treat as no_results.
    target_alnum = re.sub(r"[^A-Za-z0-9]", "", target).upper()

    def _alnum(s: str) -> str:
        return re.sub(r"[^A-Za-z0-9]", "", s or "").upper()

    def _fuzzy_match(returned_mpn: str) -> bool:
        if not target_alnum:
            return False
        r = _alnum(returned_mpn)
        if not r:
            return False
        return target_alnum in r or r in target_alnum

    def _stock_of(d: dict) -> int:
        try:
            return int(d.get("stock_now_qty") or 0)
        except (TypeError, ValueError):
            return 0

    if channel == "LCSC":
        # variants is list of variant_rec dicts with inline `extracted`
        ok_variants = [v for v in variants if v.get("status") == "ok" and v.get("extracted")]
        if not ok_variants:
            return None, len(variants)
        exact = [v for v in ok_variants
                 if str(v["extracted"].get("manufacturer_part_number", "")).strip().lower() == target]
        if exact:
            pool = exact
        else:
            fuzzy = [v for v in ok_variants
                     if _fuzzy_match(v["extracted"].get("manufacturer_part_number", ""))]
            if not fuzzy:
                return None, len(variants)
            pool = fuzzy
        best = max(pool, key=lambda v: _stock_of(v["extracted"]))
        return best["extracted"], len(variants)

    if channel == "FUTURE":
        # variants is list of lifted summary dicts; full extracted lives in
        # the per-variant subfolder's <MPN>.json
        candidates: list[dict] = []
        for v in variants:
            mpn = v.get("manufacturer_part_number") or ""
            sub = v.get("subfolder")
            if not sub:
                continue
            sub_json = out_dir / sub / f"{_safe_folder(mpn)}.json"
            if not sub_json.exists():
                continue
            try:
                vr = json.loads(sub_json.read_text(encoding="utf-8"))
            except json.JSONDecodeError:
                continue
            ex = vr.get("extracted") if vr.get("status") == "ok" else None
            if ex:
                candidates.append(ex)
        if not candidates:
            return None, len(variants)
        exact = [ex for ex in candidates
                 if str(ex.get("manufacturer_part_number", "")).strip().lower() == target]
        if exact:
            pool = exact
        else:
            fuzzy = [ex for ex in candidates
                     if _fuzzy_match(ex.get("manufacturer_part_number", ""))]
            if not fuzzy:
                return None, len(variants)
            pool = fuzzy
        best = max(pool, key=_stock_of)
        return best, len(variants)

    return None, 0


# --- index row + compare row -----------------------------------------------


# --- helpers added for the v2 (warehouse-exploded, API-aligned) schema -----

def _vendor_sku(source: str, ex: dict) -> str:
    """Per-source vendor SKU lookup. Returns "" when the source has no
    canonical SKU concept (Future, HQEW, Rochester)."""
    keys = {
        "LCSC":      ["lcsc_part_number"],
        "DIGIKEY":   ["digikey_part_number"],
        "RSONLINE":  ["rs_stock_no"],
        "ICKEY":     ["sku_id", "product_id"],
        "ONEYAC":    ["product_id"],
    }.get(source, [])
    for k in keys:
        v = ex.get(k)
        if v not in (None, ""):
            return str(v)
    return ""


_LEAD_TIME_PATTERNS = [
    (re.compile(r"Factory Lead Time:\s*(\d+)\s*Weeks?", re.I),  lambda n: int(n) * 7),
    (re.compile(r"Factory Lead Time:\s*(\d+)\s*Days?",  re.I),  int),
    (re.compile(r"原厂(?:标准)?交货期\s*(\d+)\s*周"),               lambda n: int(n) * 7),
    (re.compile(r"原厂(?:标准)?交货期\s*(\d+)\s*[天日]"),            int),
    (re.compile(r"lead\s*(\d+)\s*天"),                            int),
    (re.compile(r"(\d+)\s*天数"),                                 int),
    (re.compile(r"(\d+)\s*工作日"),                                int),
    (re.compile(r"(\d+)\s*Weeks?", re.I),                         lambda n: int(n) * 7),
    # ONEYAC shorthand: 交期 16W (weeks). The W is uppercase and not followed
    # by "eek". Anchor on word boundary so we don't catch e.g. "WROOM".
    (re.compile(r"交期\s*(\d+)\s*W\b", re.I),                     lambda n: int(n) * 7),
    (re.compile(r"交期\s*(\d+)\s*[天日]"),                          int),
]


def _parse_lead_time_days(ship_text: str, ex: dict, row: dict) -> int | None:
    """Extract lead-time-in-days from a warehouse row's ship_text.

    Channel-native wording varies — try the common patterns. Returns None when
    no pattern matches (the cell stays empty in the CSV).
    """
    txt = ship_text or ""
    for pat, fn in _LEAD_TIME_PATTERNS:
        m = pat.search(txt)
        if m:
            try:
                return fn(m.group(1))
            except (ValueError, TypeError):
                continue
    # Future Factory-Stock rows: fall back to `site_factory_lead_time`
    if row.get("label") == "Factory Stock":
        flt = ex.get("site_factory_lead_time") or ""
        for pat, fn in _LEAD_TIME_PATTERNS:
            m = pat.search(str(flt))
            if m:
                try:
                    return fn(m.group(1))
                except (ValueError, TypeError):
                    continue
    return None


def _warehouse_rows(source: str, ex: dict) -> list[dict]:
    """Explode `extracted.stock_breakdown` into per-warehouse dicts.

    Returns [] when the breakdown is missing or empty (caller emits a single
    fallback row with warehouse-level columns blank in that case).
    """
    breakdown = ex.get("stock_breakdown") or []
    if not breakdown:
        return []
    out: list[dict] = []
    for i, row in enumerate(breakdown, 1):
        out.append({
            "warehouse":      row.get("warehouse") or "",
            "warehouse_idx":  i,
            "ships_from":     row.get("ships_from") or "",
            "stockpool_qty":  row.get("quantity"),
            "ship_text":      row.get("ship_text") or "",
            "lead_time_days": _parse_lead_time_days(row.get("ship_text") or "", ex, row),
            "moq":            row.get("moq") or ex.get("min_order_qty") or ex.get("min_buy_number"),
        })
    return out


def derive_price_summary(prices: list[dict]) -> dict:
    """Summarise a tier-price list with the same scalars the API batch driver emits.

    Returned keys:
      • min_break_qty     — smallest qty break
      • price_at_min_qty  — unit price at qty=1 if present, else at smallest break
      • max_break_qty     — largest qty break
      • price_at_max_qty  — unit price at largest break (cheapest unit price)
      • num_price_tiers   — count
    """
    if not prices:
        return {"min_break_qty": None, "price_at_min_qty": None,
                "max_break_qty": None, "price_at_max_qty": None,
                "num_price_tiers": 0}
    tiers = [t for t in prices if isinstance(t, dict)]

    def _num(t):
        for k in ("unit_price_float", "unit_price_cny", "unit_price"):
            v = t.get(k)
            if v is not None and v != "":
                if isinstance(v, (int, float)):
                    return float(v)
                parsed = _parse_price_str(v)
                if parsed is not None:
                    return parsed
        return None

    by_qty = sorted([t for t in tiers if t.get("min_qty") is not None],
                    key=lambda t: t.get("min_qty"))
    min_break = by_qty[0] if by_qty else None
    largest = by_qty[-1] if by_qty else None
    one_break = next((t for t in by_qty if t.get("min_qty") == 1), None)
    return {
        "min_break_qty":    (min_break.get("min_qty") if min_break else None),
        "price_at_min_qty": (_num(one_break) if one_break else (_num(min_break) if min_break else None)),
        "max_break_qty":    (largest.get("min_qty") if largest else None),
        "price_at_max_qty": (_num(largest) if largest else None),
        "num_price_tiers":  len(tiers),
    }


def make_index_rows(
    input_mpn: str,
    expected_mfr: str,
    source: str,
    sub_result: dict,
    record: dict | None,
    extracted: dict | None,
    num_variants: int,
    run_subdir: Path,
) -> list[dict]:
    """Build one CSV row per `extracted.stock_breakdown[]` entry (warehouse-
    exploded, v2). When `status != ok` or the breakdown is empty, returns a
    single fallback row with warehouse-level columns blank — mirroring the
    API track's emit semantics in `api/doc/batch_output_schema.md`.

    Column order is fixed in INDEX_COLUMNS (caller's responsibility).
    """
    ex = extracted or {}

    # Compose final status: subprocess outcome ∧ record content.
    # The scrapers write their JSON BEFORE the final print(), so a successful
    # scrape that crashes during console output still leaves a valid record on
    # disk. Trust the record's own status when it exists.
    sub_status = sub_result.get("status")
    rec_status = (record or {}).get("status") if record else None

    if rec_status == "ok" and extracted:
        status = "ok"
    elif rec_status in ("no_matches", "no_results", "no_listings",
                         "no_detail_anchors", "no_clickable_product"):
        status = "no_results"
    elif rec_status == "blocked":
        status = "blocked"
    elif rec_status == "exception":
        status = "exception"
    elif sub_status == "timeout":
        status = "timeout"
    elif sub_status != "ok":
        # Subprocess crashed AND no record on disk → real failure
        status = sub_status
    else:
        status = rec_status or "no_results"

    prices_summary = derive_price_summary(ex.get("prices") or [])
    returned_mfr = ex.get("manufacturer") or ""

    # Currency: LCSC uses CNY implicitly when stock_now_qty present; Digikey
    # has currency in price tiers; Future has it.
    currency = ex.get("currency")
    if not currency and source == "LCSC" and ex.get("unit_price_cny") is not None:
        currency = "CNY"

    error = sub_result.get("error") or ""
    if not error and rec_status == "exception":
        error = (record or {}).get("error", "")
    if not error and rec_status == "blocked":
        error = (record or {}).get("blocker", "blocked")

    # Fields that are constant across the warehouse rows of one (MPN × source).
    base = {
        "input_mpn":         input_mpn,
        "expected_mfr":      expected_mfr,
        # Internal short code (LCSC / HQEW / ...). Transformed to the
        # bilingual SOURCE_LABEL only at CSV/XLSX write time so downstream
        # bookkeeping that pivots on `source` (failures dedup, per-source
        # summary stats, --only flag, folder names) keeps using the enum.
        "source":            source,
        "status":            status,
        "returned_mpn":      ex.get("manufacturer_part_number") or "",
        "vendor_sku":        _vendor_sku(source, ex),
        "returned_mfr":      returned_mfr,
        "mfr_match":         _mfr_match(expected_mfr, returned_mfr),
        "min_break_qty":     prices_summary["min_break_qty"],
        "price_at_min_qty":  prices_summary["price_at_min_qty"],
        "max_break_qty":     prices_summary["max_break_qty"],
        "price_at_max_qty":  prices_summary["price_at_max_qty"],
        "num_price_tiers":   prices_summary["num_price_tiers"],
        "currency":          currency or "",
        "datasheet_url":     ex.get("datasheet_url") or "",
        "run_subdir":        str(run_subdir.relative_to(PROJECT_ROOT)).replace("\\", "/"),
        "error":             (error or "")[:300],
        "elapsed_sec":       sub_result.get("elapsed_sec"),
        "num_variants":      num_variants,
    }

    wh_rows = _warehouse_rows(source, ex) if status == "ok" else []
    if not wh_rows:
        return [{**base,
                 "warehouse": "", "warehouse_idx": None, "ships_from": "",
                 "stockpool_qty": None, "ship_text": "",
                 "lead_time_days": None, "moq": None}]
    return [{**base, **w} for w in wh_rows]


INDEX_COLUMNS = [
    # 24 cols mirroring api/doc/batch_output_schema.md
    "input_mpn", "expected_mfr", "source", "status",
    "returned_mpn", "vendor_sku", "returned_mfr", "mfr_match",
    "warehouse", "warehouse_idx", "ships_from",
    "stockpool_qty", "ship_text", "lead_time_days", "moq",
    "min_break_qty", "price_at_min_qty", "max_break_qty", "price_at_max_qty",
    "num_price_tiers", "currency", "datasheet_url",
    "run_subdir", "error",
    # 2 scraper-only extras at the end
    "elapsed_sec", "num_variants",
]


CHANNEL_ORDER = ["LCSC", "DIGIKEY", "HQEW", "FUTURE", "RSONLINE", "ONEYAC", "ICKEY", "ROCHESTER"]

# Bilingual label written into the CSV `source` column. Internal code paths
# (folder names, CHANNELS dict keys, --only flag values) still use the
# short English enum; only the visible CSV cell gets the suffix so a human
# reviewing the file can tell which Chinese site it is at a glance.
SOURCE_LABEL = {
    "LCSC":      "LCSC_立创商城",
    "DIGIKEY":   "DIGIKEY_得捷电子",
    "HQEW":      "HQEW_华强电子网",
    "FUTURE":    "FUTURE_Future_Electronics",
    "RSONLINE":  "RSONLINE_RS欧时",
    "ONEYAC":    "ONEYAC_唯样商城",
    "ICKEY":     "ICKEY_云汉芯城",
    "ROCHESTER": "ROCHESTER_Rochester_Electronics",
}


# --- writers ---------------------------------------------------------------


def _cell_value_for_excel(v):
    if v is None:
        return None
    if isinstance(v, (str, int, float, bool)):
        return v
    return str(v)


def write_csv(rows: list[dict], columns: list[str], path: Path) -> None:
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=columns)
        w.writeheader()
        for r in rows:
            w.writerow({k: ("" if r.get(k) is None else r.get(k)) for k in columns})


def write_xlsx(rows: list[dict], columns: list[str], path: Path, sheet_name: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name[:31] or "Sheet1"
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="FFE2E2E2", end_color="FFE2E2E2", fill_type="solid")
    for col_idx, col in enumerate(columns, 1):
        c = ws.cell(row=1, column=col_idx, value=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(vertical="center")
    for row_idx, r in enumerate(rows, 2):
        for col_idx, col in enumerate(columns, 1):
            ws.cell(row=row_idx, column=col_idx, value=_cell_value_for_excel(r.get(col)))
    ws.freeze_panes = "A2"
    for col_idx, col in enumerate(columns, 1):
        sample = [str(r.get(col)) for r in rows if r.get(col) is not None][:300]
        max_len = max([len(col)] + [len(s) for s in sample] + [0])
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 60)
    wb.save(path)


def write_batch_input_csv(chips: list[dict], path: Path) -> None:
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(["xlsx_row", "input_mpn", "expected_mfr"])
        for c in chips:
            w.writerow([c["row"], c["input_mpn"], c["expected_mfr"]])


def write_failures(rows: list[dict], path: Path) -> None:
    # Failure rows for non-ok cells. v2 schema explodes by warehouse, but a
    # failed (chip × source) emits exactly one row (warehouse columns blank),
    # so dedup by (input_mpn, source) is not needed here — every row is unique.
    failures = [r for r in rows if r["status"] != "ok"]
    lines = ["# Batch scraper failures", ""]
    if not failures:
        lines.append("_No failures._")
        path.write_text("\n".join(lines), encoding="utf-8")
        return
    lines.append(f"{len(failures)} non-ok rows (out of {len(rows)} total).")
    lines.append("")
    by_src: dict[str, list[dict]] = {}
    for r in failures:
        by_src.setdefault(r["source"], []).append(r)
    for src in CHANNEL_ORDER:
        src_rows = by_src.get(src, [])
        if not src_rows:
            continue
        lines.append(f"## {src} — {len(src_rows)} failure(s)")
        lines.append("")
        lines.append("| input_mpn | expected_mfr | status | elapsed_sec | error | run_subdir |")
        lines.append("|---|---|---|---|---|---|")
        for r in src_rows:
            err = (r.get("error") or "").replace("|", "\\|").replace("\n", " ")[:200]
            lines.append(
                f"| `{r['input_mpn']}` | {r['expected_mfr']} | {r['status']} | "
                f"{r.get('elapsed_sec','')} | {err} | `{r['run_subdir']}` |"
            )
        lines.append("")
    path.write_text("\n".join(lines), encoding="utf-8")


def write_summary_md(
    chips: list[dict],
    rows: list[dict],
    skipped: list[dict],
    channels_used: list[str],
    batch_dir: Path,
    started: datetime,
    finished: datetime,
    json_records: list[dict] | None = None,
) -> None:
    """Render batch_summary.md from the warehouse-exploded `rows` list.

    Per-source stats and rankings are computed by deduplicating to one entry
    per `(input_mpn, source)` pair — warehouse rows of the same cell share
    the same `status`, `elapsed_sec`, `vendor_sku`, etc. so it's safe to keep
    the first row in source order.
    """
    lines: list[str] = []
    elapsed = (finished - started).total_seconds()
    lines.append(f"# Batch scraper sweep — {len(chips)} chips × {len(channels_used)} channels")
    lines.append("")
    lines.append(f"- **Started:** {started.strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"- **Finished:** {finished.strftime('%Y-%m-%d %H:%M:%S')}  (elapsed {elapsed:.1f} s ≈ {elapsed/60:.1f} min)")
    lines.append(f"- **Chips processed:** {len(chips)}  ({len(skipped)} row(s) skipped from xlsx)")
    lines.append(f"- **Warehouse-row count:** {len(rows)}  (sources: {', '.join(channels_used)})")
    lines.append("")

    # Dedupe to one entry per (input_mpn, source) for cell-level stats.
    cell_rows: dict[tuple[str, str], dict] = {}
    for r in rows:
        key = (r["input_mpn"], r["source"])
        cell_rows.setdefault(key, r)
    cells = list(cell_rows.values())

    # Per-source pass rate
    per_src: dict[str, dict] = {}
    for r in cells:
        d = per_src.setdefault(r["source"], {"ok": 0, "no_results": 0,
                                             "blocked": 0, "timeout": 0,
                                             "failed_other": 0, "total": 0,
                                             "elapsed_total": 0.0})
        d["total"] += 1
        d["elapsed_total"] += float(r.get("elapsed_sec") or 0)
        s = r["status"]
        if s == "ok":
            d["ok"] += 1
        elif s == "no_results":
            d["no_results"] += 1
        elif s == "blocked":
            d["blocked"] += 1
        elif s == "timeout":
            d["timeout"] += 1
        else:
            d["failed_other"] += 1

    lines.append("## Per-source results")
    lines.append("")
    lines.append("| Source | OK | No results | Blocked | Timeout | Other fail | Total | OK % | Mean s/MPN |")
    lines.append("|---|---|---|---|---|---|---|---|---|")
    for src in channels_used:
        d = per_src.get(src, {"ok": 0, "no_results": 0, "blocked": 0,
                              "timeout": 0, "failed_other": 0, "total": 0,
                              "elapsed_total": 0.0})
        pct = (100.0 * d["ok"] / d["total"]) if d["total"] else 0.0
        mean = (d["elapsed_total"] / d["total"]) if d["total"] else 0.0
        lines.append(
            f"| {src} | {d['ok']} | {d['no_results']} | {d['blocked']} | "
            f"{d['timeout']} | {d['failed_other']} | {d['total']} | "
            f"{pct:.1f}% | {mean:.1f} |"
        )
    lines.append("")

    # Cross-source coverage per chip: how many sources returned ok for the chip
    by_chip: dict[str, set[str]] = {}
    for r in cells:
        if r["status"] == "ok":
            by_chip.setdefault(r["input_mpn"], set()).add(r["source"])
    coverage_hist: dict[int, int] = {}
    for chip in chips:
        n_ok = len(by_chip.get(chip["input_mpn"], set()))
        coverage_hist[n_ok] = coverage_hist.get(n_ok, 0) + 1
    lines.append("## Cross-source coverage per chip")
    lines.append("")
    lines.append("| # sources returning ok | # chips |")
    lines.append("|---|---|")
    for n in sorted(coverage_hist.keys(), reverse=True):
        lines.append(f"| {n} | {coverage_hist[n]} |")
    lines.append("")

    # Highlights — top 5 by chip-level stock_now_qty (from the JSON record).
    # The warehouse-exploded CSV has per-pool stockpool_qty; for ranking
    # purposes use the cell-level `extracted_best.stock_now_qty` if available.
    chip_stock: dict[tuple[str, str], int] = {}
    if json_records:
        for rec in json_records:
            ex = rec.get("extracted_best") or {}
            q = ex.get("stock_now_qty")
            if isinstance(q, int):
                chip_stock[(rec["input_mpn"], rec.get("source") or rec.get("channel"))] = q
    lines.append("## Highlights — top 5 by chip-level in-stock quantity")
    lines.append("")
    for src in channels_used:
        lines.append(f"### {src}")
        lines.append("")
        candidates = [(mpn, q) for (mpn, s), q in chip_stock.items() if s == src and q]
        if not candidates:
            lines.append("_no in-stock results_")
            lines.append("")
            continue
        candidates.sort(key=lambda t: t[1], reverse=True)
        lines.append("| input_mpn | returned_mfr | stock_now_qty | price_at_min_qty | currency |")
        lines.append("|---|---|---|---|---|")
        for mpn, q in candidates[:5]:
            r = cell_rows.get((mpn, src), {})
            price = r.get("price_at_min_qty")
            price_s = "" if price is None else str(price)
            lines.append(
                f"| `{mpn}` | {r.get('returned_mfr','')} | {q:,} | {price_s} | {r.get('currency','')} |"
            )
        lines.append("")

    # Manufacturer mismatches per source
    lines.append("## Manufacturer mismatches (returned_mfr ≠ expected_mfr after normalization)")
    lines.append("")
    for src in channels_used:
        mismatches = [r for r in cells
                      if r["source"] == src and r["status"] == "ok"
                      and r["returned_mfr"] and r["mfr_match"] is False]
        if not mismatches:
            lines.append(f"- **{src}:** none")
            continue
        lines.append(f"- **{src}: {len(mismatches)} chip(s)**")
        lines.append("")
        lines.append("  | input_mpn | expected_mfr | returned_mfr |")
        lines.append("  |---|---|---|")
        for r in mismatches[:20]:
            lines.append(f"  | `{r['input_mpn']}` | {r['expected_mfr']} | {r['returned_mfr']} |")
        if len(mismatches) > 20:
            lines.append(f"  | …and {len(mismatches)-20} more (see batch_index.csv) |  |  |")
        lines.append("")

    if skipped:
        lines.append("## Skipped xlsx rows")
        lines.append("")
        lines.append("| row | raw_mpn | raw_mfr | reason |")
        lines.append("|---|---|---|---|")
        for s in skipped:
            lines.append(f"| {s['row']} | `{s['raw_mpn']}` | {s['raw_mfr']} | {s['reason']} |")
        lines.append("")

    lines.append("## Files in this batch folder")
    lines.append("")
    for name, what in [
        ("batch_summary.md", "this file"),
        ("batch_index.csv / .xlsx", "warehouse-exploded — one row per (MPN × source × warehouse). See scraper/doc/batch_output_schema.md."),
        ("batch_index.json", "machine-readable per-(MPN × source) records (full `record` + `extracted_best`)"),
        ("batch_input.csv", "verbatim (MPN, expected_mfr) input rows"),
        ("failures.md", "non-ok rows grouped by source"),
        ("Test_<sanitized_mpn>_<SOURCE>/", "per-MPN-per-source run folder"),
    ]:
        lines.append(f"- `{name}` — {what}")
    lines.append("")

    (batch_dir / "batch_summary.md").write_text("\n".join(lines), encoding="utf-8")


# --- main ------------------------------------------------------------------


def _parse_only(s: str | None) -> list[str]:
    if not s:
        return CHANNEL_ORDER.copy()
    requested = [tok.strip().upper() for tok in s.split(",") if tok.strip()]
    unknown = [c for c in requested if c not in CHANNELS]
    if unknown:
        raise SystemExit(f"Unknown channel(s): {unknown}. Choose from {CHANNEL_ORDER}.")
    return [c for c in CHANNEL_ORDER if c in requested]


def process_one_channel(
    channel: str,
    mpn: str,
    mfr: str,
    batch_dir: Path,
    resume_mode: bool,
) -> tuple[str, dict, dict]:
    """Run (or resume) one (MPN × channel) call. Returns (channel, index_row, record_log).

    Pure per-call work — no shared state, no printing. Safe to dispatch from a
    thread pool because the underlying network I/O happens inside a subprocess
    (truly parallel, not GIL-blocked).
    """
    safe = _safe_folder(mpn)
    run_dir = batch_dir / f"Test_{safe}_{channel}"
    run_dir.mkdir(parents=True, exist_ok=True)

    if resume_mode and (run_dir / f"{safe}.json").exists():
        record = load_record(run_dir, mpn)
        extracted, n_var = pick_best_extracted(record or {}, channel, run_dir, mpn)
        sub_result = {"status": "ok", "elapsed_sec": None, "error": ""}
        rows = make_index_rows(mpn, mfr, channel, sub_result, record, extracted, n_var, run_dir)
        record_log = {
            "source": channel, "input_mpn": mpn, "expected_mfr": mfr,
            "elapsed_sec": None, "subprocess_status": "resume_skipped",
            "record": record, "extracted_best": extracted,
        }
        return channel, rows, record_log

    sub_result = run_subprocess(channel, mpn, run_dir)
    record = load_record(run_dir, mpn)
    extracted, n_var = pick_best_extracted(record or {}, channel, run_dir, mpn)
    rows = make_index_rows(mpn, mfr, channel, sub_result, record, extracted, n_var, run_dir)
    record_log = {
        "source": channel, "input_mpn": mpn, "expected_mfr": mfr,
        "elapsed_sec": sub_result.get("elapsed_sec"),
        "subprocess_status": sub_result.get("status"),
        "stdout_tail": sub_result.get("stdout_tail"),
        "stderr_tail": sub_result.get("stderr_tail"),
        "record": record,
        "extracted_best": extracted,
        "error": sub_result.get("error"),
    }
    return channel, rows, record_log


def main(argv: list[str]) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX,
                        help=f"Path to the chip-list xlsx (default: {DEFAULT_XLSX}).")
    parser.add_argument("--mpns", type=str, default=None,
                        help="Semicolon-separated MPN list — overrides --xlsx entirely. "
                             "(Semicolon, not comma, because some MPNs contain commas — "
                             "e.g. `BT168GW,115`.) Optional `MPN:expected_mfr` syntax "
                             "per entry: 'STM32G030F6P6:STM;BT168GW,115:WEEN'. "
                             "NB: MPNs containing `:` (e.g. typo'd `BTA206X-800CT:127`) "
                             "will be wrongly chopped by the `:` separator — use --mpns-file "
                             "with tab separator for those.")
    parser.add_argument("--mpns-file", type=str, default=None,
                        help="Tab-separated file: one MPN per line, format 'MPN<TAB>MFR' "
                             "(MFR optional). Tab is the separator because some real MPNs "
                             "contain ':' . Lines starting with `#` are comments. Overrides "
                             "--xlsx and takes precedence over --mpns when both are given.")
    parser.add_argument("--limit", type=int, default=None,
                        help="Process only the first N valid MPNs (dry-run aid).")
    parser.add_argument("--only", type=str, default=None,
                        help="Comma-separated subset of channels (e.g. LCSC,HQEW).")
    parser.add_argument("--throttle", type=float, default=THROTTLE_SECONDS,
                        help=f"Sleep between chips (s). Default {THROTTLE_SECONDS}. "
                             "With parallel channels, this is a politeness gap between chips, "
                             "not between every channel call.")
    parser.add_argument("--resume", action="store_true",
                        help="Reuse the most recent BatchTest_* folder and skip rows whose "
                             "per-channel <safe_mpn>.json already exists.")
    parser.add_argument("--sequential", action="store_true",
                        help="Run channels one at a time per chip (old behavior). Default is "
                             "to fan out all selected channels concurrently per chip — each "
                             "channel hits a different domain so there's no per-vendor "
                             "rate-limit collision. Wallclock per chip drops from sum(channels) "
                             "to max(channels), typically a 50–60%% speedup.")
    parser.add_argument("--with-bom2buy", dest="with_bom2buy", action="store_true",
                        default=True,
                        help="After the main sweep, run scrape_bom2buy.py for all MPNs and "
                             "merge into batch_index. Default ON. Requires Opera fully closed "
                             "and a captcha-cleared Opera session. If the captcha session is "
                             "expired (script exit code 3) the batch still completes; the user "
                             "is told to refresh Opera + re-run scrape_bom2buy.py manually.")
    parser.add_argument("--no-bom2buy", dest="with_bom2buy", action="store_false",
                        help="Skip the bom2buy post-step (e.g. for headless CI runs that can't "
                             "drive Opera).")
    args = parser.parse_args(argv[1:])

    channels_used = _parse_only(args.only)
    if args.mpns_file:
        # Tab-separated file: one MPN per line, format `MPN\tMFR` (MFR optional).
        # Tab is the separator because some real MPNs contain ":" (e.g. typo'd
        # `BTA206X-800CT:127`). Lines starting with `#` are comments.
        chips = []
        skipped = []
        for i, line in enumerate(Path(args.mpns_file).read_text(encoding="utf-8").splitlines()):
            line = line.rstrip("\r\n")
            if not line.strip() or line.lstrip().startswith("#"):
                continue
            if "\t" in line:
                mpn_s, mfr_s = line.split("\t", 1)
            else:
                mpn_s, mfr_s = line, ""
            chips.append({"row": i + 1, "input_mpn": mpn_s.strip(), "expected_mfr": mfr_s.strip()})
        print(f"Loaded {len(chips)} chip rows from --mpns-file (xlsx ignored)")
    elif args.mpns:
        chips = []
        skipped = []
        for i, raw in enumerate(args.mpns.split(";")):
            entry = raw.strip()
            if not entry:
                continue
            if ":" in entry:
                mpn_s, mfr_s = entry.split(":", 1)
                mpn_s, mfr_s = mpn_s.strip(), mfr_s.strip()
            else:
                mpn_s, mfr_s = entry, ""
            chips.append({"row": i + 1, "input_mpn": mpn_s, "expected_mfr": mfr_s})
        print(f"Loaded {len(chips)} chip rows from --mpns flag (xlsx ignored)")
    else:
        chips, skipped = load_chip_list(args.xlsx)
        print(f"Loaded {len(chips)} chip rows ({len(skipped)} skipped) from {args.xlsx.name}")
    if args.limit:
        chips = chips[: args.limit]
    print(f"Channels: {channels_used}")

    # Batch folder selection
    if args.resume:
        existing = sorted(SCRAPER_TEST_ROOT.glob("BatchTest_*"))
        if not existing:
            print("[resume] no existing BatchTest_* folder; creating new one.")
            args.resume = False
    if not args.resume:
        now = datetime.now()
        batch_dir = SCRAPER_TEST_ROOT / f"BatchTest_{now.strftime('%Y%m%d')}_{now.strftime('%H_%M_%S')}"
    else:
        batch_dir = sorted(SCRAPER_TEST_ROOT.glob("BatchTest_*"))[-1]
        print(f"[resume] using existing batch folder {batch_dir.name}")
    batch_dir.mkdir(parents=True, exist_ok=True)
    print(f"Batch folder: {batch_dir}")

    write_batch_input_csv(chips, batch_dir / "batch_input.csv")

    started = datetime.now(timezone.utc)
    index_rows: list[dict] = []
    all_records: list[dict] = []

    parallel = (not args.sequential) and len(channels_used) > 1
    mode_label = "parallel" if parallel else "sequential"
    print(f"Channel dispatch: {mode_label} ({len(channels_used)} channel(s) per chip)")

    for i, chip in enumerate(chips, 1):
        mpn = chip["input_mpn"]
        mfr = chip["expected_mfr"]
        chip_started = time.time()
        print(f"[{i:>3}/{len(chips)}] {mpn}  (expected {mfr})")

        if parallel:
            # Fan out: one subprocess per channel, all running concurrently.
            # Each channel hits a different domain, so there's no per-vendor
            # rate-limit collision. The chip's wallclock is max(channel times)
            # instead of sum, which is the primary speedup.
            with ThreadPoolExecutor(max_workers=len(channels_used)) as ex:
                futures = [
                    ex.submit(process_one_channel, ch, mpn, mfr, batch_dir, args.resume)
                    for ch in channels_used
                ]
                channel_results = [f.result() for f in futures]
            # Re-order back to channels_used order for deterministic output
            channel_results.sort(key=lambda t: channels_used.index(t[0]))
        else:
            channel_results = [
                process_one_channel(ch, mpn, mfr, batch_dir, args.resume)
                for ch in channels_used
            ]

        for channel, rows, record_log in channel_results:
            index_rows.extend(rows)
            all_records.append(record_log)
            # Console line: chip-level stock from the JSON record (rows are
            # warehouse-exploded so they don't carry a chip-level aggregate).
            ex_best = record_log.get("extracted_best") or {}
            qty = ex_best.get("stock_now_qty") if ex_best else None
            qty_s = f"{qty:,}" if isinstance(qty, int) else (str(qty) if qty is not None else "?")
            status = rows[0]["status"] if rows else "unknown"
            n_var = rows[0].get("num_variants", 0) if rows else 0
            n_wh = sum(1 for r in rows if r.get("warehouse_idx"))
            elapsed = rows[0].get("elapsed_sec") if rows else None
            if record_log.get("subprocess_status") == "resume_skipped":
                print(f"      {channel}: [resume] {status}  qty={qty_s}  variants={n_var}  warehouses={n_wh}")
            else:
                el_s = f"{elapsed:.1f}" if isinstance(elapsed, (int, float)) else "?"
                print(f"      {channel}: {status}  qty={qty_s}  variants={n_var}  warehouses={n_wh}  ({el_s} s)")

        if parallel:
            chip_wall = time.time() - chip_started
            print(f"      chip wallclock: {chip_wall:.1f} s")

        # Politeness gap between chips. With parallel channels this fires
        # once per chip (was once per channel call in sequential mode).
        if args.throttle > 0 and i < len(chips):
            time.sleep(args.throttle)

    finished = datetime.now(timezone.utc)

    # Outputs.
    # The CSV / XLSX `source` column carries the bilingual SOURCE_LABEL
    # ("LCSC_立创商城" etc.) — a human reviewing the file can tell the
    # site at a glance. In-memory rows keep the short enum for the dedup /
    # summary code paths.
    csv_rows = [{**r, "source": SOURCE_LABEL.get(r["source"], r["source"])}
                for r in index_rows]
    write_csv(csv_rows, INDEX_COLUMNS, batch_dir / "batch_index.csv")
    write_xlsx(csv_rows, INDEX_COLUMNS, batch_dir / "batch_index.xlsx", "batch_index")
    (batch_dir / "batch_index.json").write_text(
        json.dumps(all_records, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8",
    )
    write_failures(index_rows, batch_dir / "failures.md")
    write_summary_md(
        chips, index_rows, skipped, channels_used,
        batch_dir, started.astimezone(), finished.astimezone(),
        json_records=all_records,
    )

    n_cells = len({(r["input_mpn"], r["source"]) for r in index_rows})
    print(f"\nDone (main sweep). {n_cells} cells × ~{len(index_rows)/max(n_cells,1):.1f} warehouse rows/cell = {len(index_rows)} index rows.")
    print(f"Wrote: {batch_dir}")

    # ─────────────── bom2buy post-step (default ON) ────────────────
    # bom2buy is the 9th source but can't run alongside other channels
    # (Opera-only, exclusive user-data-dir lock). We invoke scrape_bom2buy.py
    # AFTER the main sweep against the same batch folder, then re-merge.
    # If the captcha session is expired (exit code 3), we print a clear
    # user-action prompt and STILL complete the batch — bom2buy can be
    # backfilled manually.
    if args.with_bom2buy:
        scripts_dir = PROJECT_ROOT / "scraper" / "scripts"
        bom2buy_script = scripts_dir / "scrape_bom2buy.py"
        merge_script = scripts_dir / "_merge_bom2buy_into_batch.py"
        if not bom2buy_script.exists():
            print("[bom2buy] scrape_bom2buy.py not found — skipping post-step")
        else:
            # Tab-separated MPN file for the scraper (covers MPNs with `:` like
            # the typo'd `BTA206X-800CT:127`).
            mpns_file = batch_dir / ".bom2buy_input.tsv"
            with open(mpns_file, "w", encoding="utf-8", newline="\n") as fp:
                fp.write("# auto-generated by batch_scraper_test.py — full chip list for bom2buy post-step\n")
                for c in chips:
                    fp.write(f"{c['input_mpn']}\t{c['expected_mfr']}\n" if c['expected_mfr']
                             else f"{c['input_mpn']}\n")
            print(f"\n[bom2buy] starting post-step (--with-bom2buy default ON); {len(chips)} MPNs")
            try:
                result = subprocess.run(
                    [sys.executable, str(bom2buy_script),
                     "--mpns-file", str(mpns_file),
                     "--out", str(batch_dir)],
                    check=False, timeout=60 * 60,  # 1h hard cap
                )
                rc = result.returncode
            except subprocess.TimeoutExpired:
                rc = -1
                print("[bom2buy] HARD TIMEOUT (1 h) — partial result; manual retry possible")
            except Exception as e:
                rc = -2
                print(f"[bom2buy] launch failed: {e} — skipping merge")
            if rc == 0:
                print("[bom2buy] post-step ok")
            elif rc == 3:
                print(
                    "\n" + "=" * 70 + "\n"
                    "[bom2buy] ⚠ SESSION EXPIRED — bom2buy was skipped.\n"
                    "  To backfill bom2buy for this batch, AFTER the batch finishes:\n"
                    "  1. Open Opera, navigate to https://www.bom2buy.com/ and solve the\n"
                    "     IconCaptcha (just click & complete it once).\n"
                    "  2. FULLY close Opera (Task Manager → kill all opera.exe if needed).\n"
                    "  3. Run:\n"
                    f"       .venv/Scripts/python.exe scraper/scripts/scrape_bom2buy.py \\\n"
                    f"           --mpns-file {mpns_file.relative_to(PROJECT_ROOT)} \\\n"
                    f"           --out {batch_dir.relative_to(PROJECT_ROOT)}\n"
                    f"       .venv/Scripts/python.exe scraper/scripts/_merge_bom2buy_into_batch.py \\\n"
                    f"           {batch_dir.relative_to(PROJECT_ROOT)}\n"
                    "  Until then, the batch is complete EXCEPT for bom2buy rows.\n"
                    + "=" * 70
                )
            else:
                print(f"[bom2buy] post-step exit code {rc} — proceeding with whatever cells were scraped")
            # Run merge regardless of bom2buy exit code — it picks up whichever
            # cell folders exist on disk. If zero cells exist, merge adds zero
            # rows (safe).
            if merge_script.exists():
                try:
                    subprocess.run(
                        [sys.executable, str(merge_script), str(batch_dir)],
                        check=False, timeout=120,
                    )
                except Exception as e:
                    print(f"[bom2buy] merge step failed: {e}")
            else:
                print("[bom2buy] _merge_bom2buy_into_batch.py not found — batch_index.csv has no bom2buy rows")

    # Refresh scraper/README.md status block so bare-shell runs (no Claude Code
    # session, hence no PostToolUse hook) also keep the doc current. Best-
    # effort: never block on it.
    regen = PROJECT_ROOT / "scraper" / "scripts" / "_update_readme_status.py"
    if regen.exists():
        try:
            subprocess.run(
                [sys.executable, str(regen)],
                timeout=10,
                check=False,
                capture_output=True,
            )
            print(f"Refreshed: scraper/README.md status block")
        except (subprocess.TimeoutExpired, OSError):
            pass

    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
