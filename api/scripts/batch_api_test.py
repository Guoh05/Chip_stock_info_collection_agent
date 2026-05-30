"""Batch-run every chip in ref/Chip_DataSource_Master.xlsx through the
distributor APIs (Mouser + Digikey + Element14 + Arrow) and produce a
warehouse-granular index of where each chip is buyable.

For each (input_mpn, expected_mfr) row, the four sources run **in parallel
within the chip** (one thread per source); chips themselves are processed
serially so per-source rate limits remain trivially satisfied. For each
(chip × source):
  - call the source's `call_api`
  - extract the "best variant" record (per-MPN folder, raw response, summary)
  - **explode** that variant's `stock_breakdown[]` into one batch_index row
    per warehouse.

Per-chip wall clock is dominated by the slowest source (typically Digikey
~3–9 s with OAuth+search) instead of the serial sum of all four. Each
source's own print output may interleave; the per-source summary lines
emitted by this driver are serialized via a print lock.

Outputs land under `<env_root>/api/BatchTest_<YYYYMMDD>_<HH_MM_SS>/`,
where `<env_root>` is `test/` (default) or `production/` (with --env prod):
  - batch_summary.md                 — TL;DR + per-source pass rate + highlights
  - batch_index.csv / .xlsx          — long form (one row per MPN × source × warehouse)
  - batch_index.json                 — same data, machine-readable
  - batch_input.csv                  — verbatim (MPN, expected_mfr) from xlsx
  - failures.md                      — non-ok rows with error excerpt
  - Test_<sanitized_mpn>_<SOURCE>/   — per-MPN run folders, same shape as a
                                       single-MPN call would produce.

Usage:
    .venv/Scripts/python.exe api/scripts/batch_api_test.py             # full sweep (test env)
    .venv/Scripts/python.exe api/scripts/batch_api_test.py --limit 3   # dry-run
    .venv/Scripts/python.exe api/scripts/batch_api_test.py --only MOUSER
    .venv/Scripts/python.exe api/scripts/batch_api_test.py --xlsx PATH
    .venv/Scripts/python.exe api/scripts/batch_api_test.py --env prod  # write to production/api/

The script is idempotent — each invocation creates a fresh timestamped batch
folder. Re-running does NOT overwrite previous batches.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
import threading
import time
import traceback
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

# Project paths --------------------------------------------------------------
PROJECT_ROOT = Path(__file__).resolve().parents[2]
ENV_ROOTS = {
    "test": PROJECT_ROOT / "test",
    "prod": PROJECT_ROOT / "production",
}
DEFAULT_XLSX = PROJECT_ROOT / "ref" / "Raw_chip_list_20260520.xlsx"
DEFAULT_XLSX_SHEET = "Part List Modify"
DEFAULT_XLSX_MPN_HEADER = "Manufacture Part Number"
DEFAULT_XLSX_MFR_HEADER = "Manufacture"
ENV_PATH = PROJECT_ROOT / "api" / ".env"

# Reuse the single-call api clients ------------------------------------------
sys.path.insert(0, str(PROJECT_ROOT / "api" / "scripts"))
import api_mouser    # type: ignore  # noqa: E402
import api_digikey   # type: ignore  # noqa: E402
import api_element14  # type: ignore  # noqa: E402
import api_arrow     # type: ignore  # noqa: E402
import api_lcsc      # type: ignore  # noqa: E402

THROTTLE_SECONDS = 0.3
SOURCES_ALL = ("MOUSER", "DIGIKEY", "ELEMENT14", "ARROW", "LCSC")

# Display name used in user-facing artifacts (batch_index.csv/.xlsx/.json,
# batch_summary.md, failures.md). Internally we still key by the short code
# (SOURCES_ALL / SOURCE_RUNNERS / _vendor_sku etc.) — only the output rendering
# uses the long form. Update both the mapping AND the preferred-order list in
# api/scripts/_update_readme_status.py when adding a source.
SOURCE_DISPLAY_NAME = {
    "MOUSER":    "Mouser_贸泽",
    "DIGIKEY":   "DIGIKEY_得捷电子",
    "ELEMENT14": "ELEMENT14_e络盟",
    "ARROW":     "ARROW_艾睿",
    "LCSC":      "LCSC_立创商城",
}


def _source_display(source: str) -> str:
    return SOURCE_DISPLAY_NAME.get(source, source)

# Element14 emits a "Stock level (total)" aggregate row whose quantity equals
# the sum of the per-region rows. We keep it in batch_index because it carries
# the buyer-facing canonical warehouse name + ship SLA ("e络盟 在库,下单后立即发货"),
# and document the double-count caveat — same handling as Arrow mirror rows.
ELEMENT14_AGG_LABEL = "Stock level (total)"

# Per-source minimum interval between successive calls (seconds). Element14's
# published quota is 2 req/s, 1000/day — we space its calls ≥0.6 s apart to
# stay safely under the burst limit even if call rotation changes.
# Values may be overridden via env: ELEMENT14_CALLS_PER_SECOND (or the legacy
# Calls_per_second_limit name in .env's Element14 block).
SOURCE_MIN_INTERVAL_DEFAULT = {
    "MOUSER": 0.0,
    "DIGIKEY": 0.0,
    "ELEMENT14": 0.6,
    "ARROW": 0.0,
    "LCSC": 0.0,  # quota observed = 200/day per endpoint, no per-second cap documented
}


def _element14_min_interval() -> float:
    raw = (
        os.environ.get("ELEMENT14_CALLS_PER_SECOND")
        or os.environ.get("Calls_per_second_limit")
        or ""
    ).strip()
    try:
        cps = float(raw)
        if cps > 0:
            # 10% safety margin
            return (1.0 / cps) * 1.1
    except ValueError:
        pass
    return SOURCE_MIN_INTERVAL_DEFAULT["ELEMENT14"]


# --- helpers ----------------------------------------------------------------


def _safe_folder(mpn: str) -> str:
    return re.sub(r"[^A-Za-z0-9._-]", "_", mpn) or "UNKNOWN"


def _looks_like_real_mpn(value: Any) -> bool:
    if value is None:
        return False
    s = str(value).strip()
    if len(s) < 3:
        return False
    if "缺失" in s or "TBD" in s.upper() or "TODO" in s.upper():
        return False
    if not re.search(r"[A-Za-z0-9]", s):
        return False
    return True


def _norm_mfr(s: str | None) -> str:
    if not s:
        return ""
    return re.sub(r"[^A-Z0-9]", "", s.upper())


def _mfr_match(expected: str | None, returned: str | None) -> bool:
    e = _norm_mfr(expected)
    r = _norm_mfr(returned)
    if not e or not r:
        return False
    return e in r or r in e


def _parse_price_str(price_str) -> float | None:
    if price_str is None or price_str == "":
        return None
    if isinstance(price_str, (int, float)):
        return float(price_str)
    s = re.sub(r"[^\d.,\-]", "", str(price_str).strip())
    if not s:
        return None
    if "," in s and "." not in s:
        s = s.replace(",", ".")
    else:
        s = s.replace(",", "")
    try:
        return float(s)
    except ValueError:
        return None


_LEAD_DAYS_RE = re.compile(r"lead\s+(\d+)\s*天|(\d+)\s*天")


def _parse_lead_days(ship_text: str | None) -> int | None:
    if not ship_text:
        return None
    m = _LEAD_DAYS_RE.search(ship_text)
    if not m:
        return None
    for g in m.groups():
        if g:
            try:
                return int(g)
            except ValueError:
                continue
    return None


# --- input loading ----------------------------------------------------------


def load_chip_list(
    xlsx_path: Path,
    sheet_name: str = DEFAULT_XLSX_SHEET,
    mpn_header: str = DEFAULT_XLSX_MPN_HEADER,
    mfr_header: str = DEFAULT_XLSX_MFR_HEADER,
) -> tuple[list[dict], list[dict]]:
    """Return ([{row, input_mpn, expected_mfr}, ...], skipped[]) — dedup'd by MPN.

    Source-of-truth contract (per user 2026-05-20):
      file:   ref/Raw_chip_list_20260520.xlsx
      sheet:  "Part List Modify"
      MPN:    column with header "Manufacture Part Number" (cells may have a
              leading non-breaking space U+00A0 — stripped here)
      mfr:    column with header "Manufacture" — informational only; first
              occurrence's value wins on dedup.

    Header row is row 1, data from row 2. Columns are looked up by HEADER NAME
    so they tolerate spreadsheet reshuffling. Multiple input rows with the same
    MPN (e.g. different projects ordering the same part) get collapsed to one.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Sheet {sheet_name!r} not found in {xlsx_path}; "
            f"available: {wb.sheetnames}"
        )
    ws = wb[sheet_name]

    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    mpn_col = headers.get(mpn_header)
    mfr_col = headers.get(mfr_header)
    if mpn_col is None:
        raise ValueError(
            f"Header {mpn_header!r} not found in row 1 of {sheet_name!r}; "
            f"got: {list(headers)}"
        )

    chips: list[dict] = []
    skipped: list[dict] = []
    seen_mpns: set[str] = set()
    for r in range(2, ws.max_row + 1):
        mpn_raw = ws.cell(row=r, column=mpn_col).value
        mfr_raw = ws.cell(row=r, column=mfr_col).value if mfr_col else None
        # Strip ASCII whitespace AND U+00A0 (the master file uses NBSP prefixes
        # on some MPNs like " PFS132") so dedup matches across encodings.
        mpn_s = (str(mpn_raw).replace("\xa0", " ").strip()
                 if mpn_raw is not None else "")
        mfr_s = (str(mfr_raw).replace("\xa0", " ").strip()
                 if mfr_raw is not None else "")
        if not _looks_like_real_mpn(mpn_s):
            skipped.append({"row": r, "raw_mpn": mpn_s, "raw_mfr": mfr_s,
                            "reason": "missing or non-MPN placeholder"})
            continue
        if mpn_s in seen_mpns:
            skipped.append({"row": r, "raw_mpn": mpn_s, "raw_mfr": mfr_s,
                            "reason": f"duplicate MPN (already loaded)"})
            continue
        seen_mpns.add(mpn_s)
        chips.append({"row": r, "input_mpn": mpn_s, "expected_mfr": mfr_s})
    return chips, skipped


# --- per-source best-variant selection --------------------------------------


def _pick_mouser(payload: dict, input_mpn: str) -> dict | None:
    parts = (payload.get("SearchResults") or {}).get("Parts") or []
    if not parts:
        return None
    target = input_mpn.strip().lower()
    exact = [p for p in parts
             if str(p.get("ManufacturerPartNumber", "")).strip().lower() == target]
    pool = exact or parts

    def _stock(p):
        try:
            return int(str(p.get("AvailabilityInStock") or "0").split()[0])
        except (ValueError, IndexError):
            return 0
    best_raw = max(pool, key=_stock)
    return api_mouser.normalize_part(best_raw, input_mpn)


def _pick_digikey(payload: dict, input_mpn: str) -> dict | None:
    exact = payload.get("ExactMatches") or []
    products = payload.get("Products") or []
    candidates: list[dict] = []
    seen: set[str] = set()
    for p in exact + products:
        mpn = p.get("ManufacturerProductNumber") or ""
        if mpn and mpn not in seen:
            candidates.append(p)
            seen.add(mpn)
    if not candidates:
        return None
    target = input_mpn.strip().lower()
    match = next(
        (p for p in candidates
         if (p.get("ManufacturerProductNumber") or "").strip().lower() == target),
        None,
    )
    chosen = match or candidates[0]
    return api_digikey.normalize_product(chosen, input_mpn)


def _pick_element14(root: dict, input_mpn: str, store_id: str) -> dict | None:
    products = (root or {}).get("products") or []
    if not products:
        return None
    target = input_mpn.strip().lower()
    candidates: list[dict] = []
    for raw in products:
        ex = api_element14.normalize_product(raw, input_mpn, store_id)
        candidates.append({"raw": raw, "ex": ex})
    # Exact match wins; else highest stock.
    exact = [c for c in candidates
             if (c["ex"].get("manufacturer_part_number") or "").strip().lower() == target]
    pool = exact or candidates
    pool.sort(key=lambda c: c["ex"].get("stock_now_qty") or 0, reverse=True)
    return pool[0]["ex"]


def _pick_arrow(parts: list[dict], input_mpn: str) -> dict | None:
    if not parts:
        return None
    target = input_mpn.strip().lower()
    candidates: list[dict] = []
    for raw in parts:
        ex = api_arrow.normalize_part(raw, input_mpn)
        candidates.append({"raw": raw, "ex": ex})
    exact = [c for c in candidates
             if (c["ex"].get("manufacturer_part_number") or "").strip().lower() == target]
    pool = exact or candidates
    pool.sort(key=lambda c: c["ex"].get("stock_now_qty") or 0, reverse=True)
    return pool[0]["ex"]


# --- per-source call wrappers (each writes the per-MPN run folder) ---------


def _write_parent_json(parent_rec: dict, run_dir: Path, input_mpn: str) -> None:
    (run_dir / f"{_safe_folder(input_mpn)}.json").write_text(
        json.dumps(parent_rec, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def run_mouser(input_mpn: str, run_dir: Path) -> dict:
    rec = api_mouser.call_api(input_mpn, run_dir)
    payload = rec.pop("raw_payload", None)
    if payload is not None:
        (run_dir / "raw_response.json").write_text(
            json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
        )
    variants_info: list[dict] = []
    chosen_ex: dict | None = None
    if rec.get("status") == "ok" and payload is not None:
        parts = (payload.get("SearchResults") or {}).get("Parts") or []
        seen: dict[str, dict] = {}
        for raw_part in parts:
            mpn = raw_part.get("ManufacturerPartNumber") or "UNKNOWN"
            ex_candidate = api_mouser.normalize_part(raw_part, input_mpn)
            prev = seen.get(mpn)
            if prev is None or (
                (ex_candidate.get("stock_now_qty") or 0)
                > (prev["extracted"].get("stock_now_qty") or 0)
            ):
                seen[mpn] = {"raw": raw_part, "extracted": ex_candidate}
        for mpn, bundle in seen.items():
            info = api_mouser.write_variant(rec, bundle["extracted"], bundle["raw"], run_dir, mpn)
            variants_info.append(info)
        chosen_ex = _pick_mouser(payload, input_mpn)
    parent_rec = dict(rec)
    parent_rec["variants_summary"] = [
        {
            "manufacturer_part_number": v["extracted"].get("manufacturer_part_number"),
            "mouser_part_number": v["extracted"].get("mouser_part_number"),
            "stock_now_qty": v["extracted"].get("stock_now_qty"),
            "stock_future_qty": v["extracted"].get("stock_future_qty"),
            "stock_future_ship_text": v["extracted"].get("stock_future_ship_text"),
            "subdir": v["folder"],
        }
        for v in variants_info
    ]
    _write_parent_json(parent_rec, run_dir, input_mpn)
    api_mouser.write_parent_summary(parent_rec, variants_info, run_dir)
    return {"record": rec, "chosen_extracted": chosen_ex, "num_variants": len(variants_info)}


def run_digikey(input_mpn: str, run_dir: Path) -> dict:
    rec = api_digikey.call_api(input_mpn, run_dir)
    payload = rec.pop("raw_payload", None)
    if payload is not None:
        (run_dir / "raw_response.json").write_text(
            json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
        )
    variants_info: list[dict] = []
    chosen_ex: dict | None = None
    if rec.get("status") == "ok" and payload is not None:
        exact = payload.get("ExactMatches") or []
        products = payload.get("Products") or []
        seen_mpns: set[str] = set()
        candidates: list[dict] = []
        for p in exact + products:
            mpn = p.get("ManufacturerProductNumber") or ""
            if mpn and mpn not in seen_mpns:
                candidates.append(p)
                seen_mpns.add(mpn)
        for product in candidates:
            mpn = product.get("ManufacturerProductNumber") or "UNKNOWN"
            extracted = api_digikey.normalize_product(product, input_mpn)
            info = api_digikey.write_variant(rec, extracted, product, run_dir, mpn)
            variants_info.append(info)
        chosen_ex = _pick_digikey(payload, input_mpn)
    parent_rec = dict(rec)
    parent_rec["variants_summary"] = [
        {
            "manufacturer_part_number": v["extracted"].get("manufacturer_part_number"),
            "digikey_part_number": v["extracted"].get("digikey_part_number"),
            "stock_now_qty": v["extracted"].get("stock_now_qty"),
            "stock_future_qty": v["extracted"].get("stock_future_qty"),
            "stock_future_ship_text": v["extracted"].get("stock_future_ship_text"),
            "subdir": v["folder"],
        }
        for v in variants_info
    ]
    _write_parent_json(parent_rec, run_dir, input_mpn)
    api_digikey.write_parent_summary(parent_rec, variants_info, run_dir)
    return {"record": rec, "chosen_extracted": chosen_ex, "num_variants": len(variants_info)}


def run_element14(input_mpn: str, run_dir: Path) -> dict:
    rec = api_element14.call_api(input_mpn, run_dir)
    payload = rec.pop("raw_payload", None)
    root = rec.pop("root", None)
    store_id = rec.get("store_id", "cn.element14.com")
    if payload is not None:
        (run_dir / "raw_response.json").write_text(
            json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
        )
    variants_info: list[dict] = []
    chosen_ex: dict | None = None
    # Element14 commonly returns multiple SKUs for the same MPN, each with a
    # different packaging (Cut Tape / Re-Reel / Full Reel). Per the "same-MPN
    # multi-SKU expansion" rule, collect every SKU whose MPN matches the chosen
    # variant's MPN so the batch_index emits one (warehouse) sub-row per SKU.
    expanded_extracteds: list[dict] = []
    if rec.get("status") == "ok" and root is not None:
        products = root.get("products") or []
        seen: dict[str, dict] = {}
        all_extracted_per_sku: list[dict] = []
        for raw in products:
            ex = api_element14.normalize_product(raw, input_mpn, store_id)
            all_extracted_per_sku.append(ex)
            mpn = ex.get("manufacturer_part_number") or "UNKNOWN"
            prev = seen.get(mpn)
            if prev is None or (
                (ex.get("stock_now_qty") or 0)
                > (prev["extracted"].get("stock_now_qty") or 0)
            ):
                seen[mpn] = {"raw": raw, "extracted": ex}
        for mpn, bundle in seen.items():
            info = api_element14.write_variant(rec, bundle["extracted"], bundle["raw"], run_dir, mpn)
            variants_info.append(info)
        chosen_ex = _pick_element14(root, input_mpn, store_id)
        # Filter by the chosen variant's MPN — only expand same-MPN multi-SKU.
        chosen_mpn = ((chosen_ex or {}).get("manufacturer_part_number") or "").strip().lower()
        if chosen_mpn:
            expanded_extracteds = [
                ex for ex in all_extracted_per_sku
                if (ex.get("manufacturer_part_number") or "").strip().lower() == chosen_mpn
            ]
    parent_rec = dict(rec)
    parent_rec["variants_summary"] = [
        {
            "manufacturer_part_number": v["extracted"].get("manufacturer_part_number"),
            "element14_sku": v["extracted"].get("element14_sku"),
            "stock_now_qty": v["extracted"].get("stock_now_qty"),
            "stock_future_qty": v["extracted"].get("stock_future_qty"),
            "stock_future_ship_text": v["extracted"].get("stock_future_ship_text"),
            "subdir": v["folder"],
        }
        for v in variants_info
    ]
    _write_parent_json(parent_rec, run_dir, input_mpn)
    api_element14.write_parent_summary(parent_rec, variants_info, run_dir)
    return {
        "record": rec,
        "chosen_extracted": chosen_ex,
        "expanded_extracteds": expanded_extracteds,
        "num_variants": len(variants_info),
    }


def run_arrow(input_mpn: str, run_dir: Path) -> dict:
    rec = api_arrow.call_api(input_mpn, run_dir)
    payload = rec.pop("raw_payload", None)
    parts = rec.pop("parts", None)
    if payload is not None:
        (run_dir / "raw_response.json").write_text(
            json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
        )
    variants_info: list[dict] = []
    chosen_ex: dict | None = None
    if rec.get("status") == "ok" and parts:
        seen: dict[str, dict] = {}
        for raw in parts:
            ex = api_arrow.normalize_part(raw, input_mpn)
            mpn = ex.get("manufacturer_part_number") or "UNKNOWN"
            prev = seen.get(mpn)
            if prev is None or (
                (ex.get("stock_now_qty") or 0)
                > (prev["extracted"].get("stock_now_qty") or 0)
            ):
                seen[mpn] = {"raw": raw, "extracted": ex}
        for mpn, bundle in seen.items():
            info = api_arrow.write_variant(rec, bundle["extracted"], bundle["raw"], run_dir, mpn)
            variants_info.append(info)
        chosen_ex = _pick_arrow(parts, input_mpn)
    parent_rec = dict(rec)
    parent_rec["variants_summary"] = [
        {
            "manufacturer_part_number": v["extracted"].get("manufacturer_part_number"),
            "arrow_item_id": v["extracted"].get("arrow_item_id"),
            "stock_now_qty": v["extracted"].get("stock_now_qty"),
            "stock_future_qty": v["extracted"].get("stock_future_qty"),
            "stock_future_ship_text": v["extracted"].get("stock_future_ship_text"),
            "subdir": v["folder"],
        }
        for v in variants_info
    ]
    _write_parent_json(parent_rec, run_dir, input_mpn)
    api_arrow.write_parent_summary(parent_rec, variants_info, run_dir)
    return {"record": rec, "chosen_extracted": chosen_ex, "num_variants": len(variants_info)}


def _pick_lcsc(data: list[dict], input_mpn: str) -> dict | None:
    if not data:
        return None
    target = input_mpn.strip().lower()
    candidates: list[dict] = []
    for raw in data:
        ex = api_lcsc.normalize_product(raw, input_mpn)
        candidates.append({"raw": raw, "ex": ex})
    exact = [c for c in candidates
             if (c["ex"].get("manufacturer_part_number") or "").strip().lower() == target]
    pool = exact or candidates
    pool.sort(key=lambda c: c["ex"].get("stock_now_qty") or 0, reverse=True)
    return pool[0]["ex"]


def run_lcsc(input_mpn: str, run_dir: Path) -> dict:
    rec = api_lcsc.call_api(input_mpn, run_dir)
    payload = rec.pop("raw_payload", None)
    if payload is not None:
        (run_dir / "raw_response.json").write_text(
            json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
        )
    variants_info: list[dict] = []
    chosen_ex: dict | None = None
    if rec.get("status") == "ok" and payload is not None:
        data = payload.get("data") or []
        # Group by exact MPN — different productIds with the same productModel
        # get the highest-stock representative.
        seen: dict[str, dict] = {}
        for raw in data:
            ex = api_lcsc.normalize_product(raw, input_mpn)
            mpn = ex.get("manufacturer_part_number") or "UNKNOWN"
            prev = seen.get(mpn)
            if prev is None or (
                (ex.get("stock_now_qty") or 0)
                > (prev["extracted"].get("stock_now_qty") or 0)
            ):
                seen[mpn] = {"raw": raw, "extracted": ex}
        for mpn, bundle in seen.items():
            info = api_lcsc.write_variant(rec, bundle["extracted"], bundle["raw"], run_dir, mpn)
            variants_info.append(info)
        chosen_ex = _pick_lcsc(data, input_mpn)
    parent_rec = dict(rec)
    parent_rec["variants_summary"] = [
        {
            "manufacturer_part_number": v["extracted"].get("manufacturer_part_number"),
            "lcsc_sku": v["extracted"].get("lcsc_sku"),
            "stock_now_qty": v["extracted"].get("stock_now_qty"),
            "stock_future_qty": v["extracted"].get("stock_future_qty"),
            "stock_future_ship_text": v["extracted"].get("stock_future_ship_text"),
            "subdir": v["folder"],
        }
        for v in variants_info
    ]
    _write_parent_json(parent_rec, run_dir, input_mpn)
    api_lcsc.write_parent_summary(parent_rec, variants_info, run_dir)
    return {"record": rec, "chosen_extracted": chosen_ex, "num_variants": len(variants_info)}


SOURCE_RUNNERS = {
    "MOUSER": run_mouser,
    "DIGIKEY": run_digikey,
    "ELEMENT14": run_element14,
    "ARROW": run_arrow,
    "LCSC": run_lcsc,
}


# --- price-tier summarization -----------------------------------------------


def derive_price_pair(tiers: list[dict] | None) -> dict:
    """Return {min_break_qty, price_at_min_qty, max_break_qty, price_at_max_qty,
    num_price_tiers}. min_qty defines the break boundary; smallest break = entry
    quantity, largest break = volume tier (typically cheapest unit price).
    """
    out = {
        "min_break_qty": None,
        "price_at_min_qty": None,
        "max_break_qty": None,
        "price_at_max_qty": None,
        "num_price_tiers": 0,
    }
    if not tiers:
        return out
    valid = [t for t in tiers if isinstance(t, dict) and t.get("min_qty") is not None]
    if not valid:
        out["num_price_tiers"] = len([t for t in tiers if isinstance(t, dict)])
        return out

    def _num(t: dict) -> float | None:
        v = t.get("unit_price_float")
        if v is not None:
            return v
        return _parse_price_str(t.get("unit_price"))
    sorted_tiers = sorted(valid, key=lambda t: t["min_qty"])
    lo = sorted_tiers[0]
    hi = sorted_tiers[-1]
    out["min_break_qty"] = lo["min_qty"]
    out["price_at_min_qty"] = _num(lo)
    out["max_break_qty"] = hi["min_qty"]
    out["price_at_max_qty"] = _num(hi)
    out["num_price_tiers"] = len(valid)
    return out


# --- index row construction -------------------------------------------------


INDEX_COLUMNS = [
    "input_mpn", "expected_mfr", "source", "status",
    "returned_mpn", "vendor_sku", "returned_mfr", "mfr_match",
    "warehouse", "warehouse_idx", "ships_from",
    "stockpool_qty", "ship_text", "lead_time_days", "moq",
    "min_break_qty", "price_at_min_qty",
    "max_break_qty", "price_at_max_qty",
    "num_price_tiers", "currency", "packaging_option",
    "datasheet_url", "run_subdir", "error",
]


def _vendor_sku(source: str, ex: dict) -> str:
    if source == "MOUSER":
        return ex.get("mouser_part_number") or ""
    if source == "DIGIKEY":
        return ex.get("digikey_part_number") or ""
    if source == "ELEMENT14":
        return ex.get("element14_sku") or ""
    if source == "ARROW":
        item = ex.get("arrow_item_id")
        return "" if item is None else str(item)
    if source == "LCSC":
        return ex.get("lcsc_sku") or ""
    return ""


def _derive_packaging_option(source: str, ex: dict, sb_row: dict,
                              site_source: dict | None) -> str:
    """Map per-row packaging info to a canonical, English label.

    DigiKey: the `Packaging — <name>` label that `api_digikey.normalize_product`
        emits on each ProductVariations[] row carries the native packaging name
        ("Tape & Reel (TR)" / "Cut Tape (CT)" / "Digi-Reel®" / …). We strip
        the "Packaging — " prefix.
    Element14: the normalize already maps unitOfMeasure + reeling → a canonical
        value ("Cut Tape" / "Full Reel" / "Re-Reel" / "Each"). Same on every
        warehouse row for a given SKU variant.
    Mouser: same — normalize derives a canonical value from Reeling + the
        Mouser PN suffix (`-TR` / `-CT`). Same on every warehouse row.
    Arrow: per-warehouse — pulled from `site_sources[i].container_type`
        (sparse; "Cut Strips" is the only observed value as of 2026-05).
    LCSC: not exposed by `search/global`. Empty.
    """
    if source == "DIGIKEY":
        label = sb_row.get("label") or ""
        if label.startswith("Packaging — "):
            return label[len("Packaging — "):].strip()
        return ""
    if source in ("ELEMENT14", "MOUSER"):
        return ex.get("packaging_option") or ""
    if source == "ARROW":
        return (site_source or {}).get("container_type") or ""
    return ""


def iter_warehouse_rows(
    input_mpn: str,
    expected_mfr: str,
    source: str,
    bundle: dict,
    run_subdir: Path,
    ex_override: dict | None = None,
) -> Iterable[dict]:
    """Yield one row per non-aggregate warehouse in the chosen variant's
    stock_breakdown[]. Element14's 'Stock level (total)' aggregate is filtered.

    `ex_override` lets the caller explode an alternate variant's
    `extracted` (used by run_element14's same-MPN multi-SKU expansion).
    """
    rec = bundle.get("record") or {}
    ex = ex_override if ex_override is not None else (bundle.get("chosen_extracted") or {})
    status = rec.get("status") or "no_results"
    returned_mpn = ex.get("manufacturer_part_number") or ""
    returned_mfr = ex.get("manufacturer") or ""
    vendor_sku = _vendor_sku(source, ex)
    top_currency = ex.get("currency") or ""
    top_moq = ex.get("min_order_qty")
    top_prices = ex.get("prices") or []
    datasheet_url = ex.get("datasheet_url") or ""
    run_subdir_str = str(run_subdir.relative_to(PROJECT_ROOT)).replace("\\", "/")

    site_sources = ex.get("site_sources") or []  # Arrow-only; index-matched 1:1
                                                 # with the first N stock_breakdown
                                                 # entries (rest are pipeline rows).

    # Top-level fallback lead time (Mouser ship_text in days, Digikey weeks*7)
    top_lead_days: int | None = None
    if source == "MOUSER":
        top_lead_days = _parse_lead_days(ex.get("site_lead_time"))
    elif source == "DIGIKEY":
        wk = ex.get("site_manufacturer_lead_weeks")
        if isinstance(wk, (int, float)) and wk > 0:
            top_lead_days = int(round(wk * 7))
    elif source == "ELEMENT14":
        top_lead_days = ex.get("site_lead_time_days")
    # Arrow's top-level lead time comes from the per-source mfr_lead_time_days;
    # we use the per-warehouse value below.

    sb = ex.get("stock_breakdown") or []
    emitted = 0
    for i, row in enumerate(sb):
        emitted += 1
        # Defaults from top-level
        moq = top_moq
        ships_from = ""
        lead_days = _parse_lead_days(row.get("ship_text")) or top_lead_days
        currency = top_currency
        price_tiers = top_prices

        # Per-warehouse overrides
        if row.get("moq") is not None:
            moq = row["moq"]
        if source == "ARROW" and i < len(site_sources):
            ss = site_sources[i]
            ships_from = ss.get("ships_from") or ""
            if ss.get("mfr_lead_time_days"):
                lead_days = ss["mfr_lead_time_days"]
            if ss.get("currency"):
                currency = ss["currency"]
            if ss.get("tiers"):
                price_tiers = ss["tiers"]

        site_source_row = site_sources[i] if (source == "ARROW" and i < len(site_sources)) else None
        packaging_option = _derive_packaging_option(source, ex, row, site_source_row)

        pp = derive_price_pair(price_tiers)
        yield {
            "input_mpn": input_mpn,
            "expected_mfr": expected_mfr,
            "source": _source_display(source),
            "status": status,
            "returned_mpn": returned_mpn,
            "vendor_sku": vendor_sku,
            "returned_mfr": returned_mfr,
            "mfr_match": _mfr_match(expected_mfr, returned_mfr),
            "warehouse": row.get("warehouse") or row.get("label") or "",
            "warehouse_idx": emitted,
            "ships_from": ships_from,
            "stockpool_qty": row.get("quantity"),
            "ship_text": row.get("ship_text") or "",
            "lead_time_days": lead_days,
            "moq": moq,
            "min_break_qty": pp["min_break_qty"],
            "price_at_min_qty": pp["price_at_min_qty"],
            "max_break_qty": pp["max_break_qty"],
            "price_at_max_qty": pp["price_at_max_qty"],
            "num_price_tiers": pp["num_price_tiers"],
            "currency": currency,
            "packaging_option": packaging_option,
            "datasheet_url": datasheet_url,
            "run_subdir": run_subdir_str,
            "error": "",
        }


def make_empty_source_row(
    input_mpn: str,
    expected_mfr: str,
    source: str,
    status: str,
    error: str,
    run_subdir: Path,
) -> dict:
    """Emit exactly one row when the (chip × source) call failed or returned
    no usable variant. Warehouse-level columns are left empty."""
    return {
        "input_mpn": input_mpn,
        "expected_mfr": expected_mfr,
        "source": _source_display(source),
        "status": status,
        "returned_mpn": "",
        "vendor_sku": "",
        "returned_mfr": "",
        "mfr_match": False,
        "warehouse": "",
        "warehouse_idx": None,
        "ships_from": "",
        "stockpool_qty": None,
        "ship_text": "",
        "lead_time_days": None,
        "moq": None,
        "min_break_qty": None,
        "price_at_min_qty": None,
        "max_break_qty": None,
        "price_at_max_qty": None,
        "num_price_tiers": 0,
        "currency": "",
        "packaging_option": "",
        "datasheet_url": "",
        "run_subdir": str(run_subdir.relative_to(PROJECT_ROOT)).replace("\\", "/"),
        "error": error or "",
    }


# --- writers ----------------------------------------------------------------


def _cell_value_for_excel(v):
    if v is None:
        return None
    if isinstance(v, (str, int, float, bool)):
        return v
    return str(v)


def write_csv(rows: list[dict], columns: list[str], path: Path) -> None:
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=columns)
        w.writeheader()
        for r in rows:
            w.writerow({k: ("" if r.get(k) is None else r.get(k)) for k in columns})


def write_xlsx(rows: list[dict], columns: list[str], path: Path, sheet_name: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name[:31] or "Sheet1"
    header_font = Font(bold=True, name="Calibri")
    header_fill = PatternFill(start_color="FFE2E2E2", end_color="FFE2E2E2", fill_type="solid")
    for col_idx, col in enumerate(columns, 1):
        c = ws.cell(row=1, column=col_idx, value=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(vertical="center")
    for row_idx, r in enumerate(rows, 2):
        for col_idx, col in enumerate(columns, 1):
            ws.cell(row=row_idx, column=col_idx, value=_cell_value_for_excel(r.get(col)))
    ws.freeze_panes = "A2"
    for col_idx, col in enumerate(columns, 1):
        max_len = max([len(col)] + [
            len(str(r.get(col))) for r in rows if r.get(col) is not None
        ][:300])
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 50)
    wb.save(path)


def write_batch_input_csv(chips: list[dict], path: Path) -> None:
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(["xlsx_row", "input_mpn", "expected_mfr"])
        for c in chips:
            w.writerow([c["row"], c["input_mpn"], c["expected_mfr"]])


def write_failures(rows: list[dict], path: Path) -> None:
    # One row per (input_mpn, source) — dedup since exploded warehouse rows
    # share the same status. Empty-source-row failures already are single rows.
    seen: set[tuple] = set()
    failures: list[dict] = []
    for r in rows:
        if r["status"] == "ok":
            continue
        key = (r["input_mpn"], r["source"])
        if key in seen:
            continue
        seen.add(key)
        failures.append(r)
    lines = ["# Batch API failures", ""]
    if not failures:
        lines.append("_No failures._")
    else:
        # Count distinct (mpn, source) attempts for the denominator
        attempted = {(r["input_mpn"], r["source"]) for r in rows}
        lines.append(f"{len(failures)} non-ok (chip × source) pairs (out of {len(attempted)} total).")
        lines.append("")
        lines.append("| input_mpn | expected_mfr | source | status | error | run_subdir |")
        lines.append("|---|---|---|---|---|---|")
        for r in failures:
            err = (r.get("error") or "").replace("|", "\\|").replace("\n", " ")[:200]
            lines.append(
                f"| `{r['input_mpn']}` | {r['expected_mfr']} | {r['source']} | "
                f"{r['status']} | {err} | `{r['run_subdir']}` |"
            )
    path.write_text("\n".join(lines), encoding="utf-8")


def write_summary_md(
    chips: list[dict],
    rows: list[dict],
    skipped: list[dict],
    batch_dir: Path,
    started: datetime,
    finished: datetime,
    sources_run: list[str],
) -> None:
    lines: list[str] = []
    elapsed = (finished - started).total_seconds()
    # `sources_run` arrives as short codes (internal IDs); rows already carry
    # display names in their "source" field. Translate once so downstream lookups
    # match.
    sources_display = [_source_display(s) for s in sources_run]
    lines.append(f"# Batch API sweep — {len(chips)} chips × {', '.join(sources_display)}")
    lines.append("")
    lines.append(f"- **Started:** {started.strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"- **Finished:** {finished.strftime('%Y-%m-%d %H:%M:%S')}  (elapsed {elapsed:.1f} s)")
    lines.append(f"- **Chips processed:** {len(chips)}  ({len(skipped)} row(s) skipped from xlsx)")
    lines.append(f"- **Total warehouse rows in batch_index:** {len(rows)}")
    lines.append("")

    # Per-source pass rate — dedup by (input_mpn, source) before counting,
    # since each chip × source contributes multiple warehouse rows on success
    # but only one on failure.
    per_source: dict[str, dict] = {}
    seen: set[tuple] = set()
    for r in rows:
        key = (r["input_mpn"], r["source"])
        if key in seen:
            continue
        seen.add(key)
        d = per_source.setdefault(r["source"], {"ok": 0, "no_results": 0,
                                                "failed_other": 0, "total": 0})
        d["total"] += 1
        if r["status"] == "ok":
            d["ok"] += 1
        elif r["status"] == "no_results":
            d["no_results"] += 1
        else:
            d["failed_other"] += 1

    lines.append("## Per-source pass rate")
    lines.append("")
    lines.append("| Source | OK | No results | Failed | Total | OK % |")
    lines.append("|---|---|---|---|---|---|")
    for src_disp in sources_display:
        d = per_source.get(src_disp, {"ok": 0, "no_results": 0, "failed_other": 0, "total": 0})
        pct = (100.0 * d["ok"] / d["total"]) if d["total"] else 0.0
        lines.append(f"| {src_disp} | {d['ok']} | {d['no_results']} | {d['failed_other']} | {d['total']} | {pct:.1f}% |")
    lines.append("")

    # Highlights — top stock-pool size per source (just the biggest single
    # warehouse, not summed; mirror rows mean naive sum can double-count).
    def _top(source_disp: str, n: int = 5):
        ok = [r for r in rows
              if r["source"] == source_disp and r["status"] == "ok"
              and (r.get("stockpool_qty") or 0) > 0]
        return sorted(ok, key=lambda r: r["stockpool_qty"], reverse=True)[:n]

    lines.append("## Highlights — top 5 single-warehouse stock pools per source")
    lines.append("")
    for src_disp in sources_display:
        lines.append(f"### {src_disp}")
        lines.append("")
        lines.append("| input_mpn | warehouse | stockpool_qty | ship_text |")
        lines.append("|---|---|---|---|")
        for r in _top(src_disp):
            lines.append(
                f"| `{r['input_mpn']}` | {r['warehouse']} | "
                f"{r['stockpool_qty']:,} | {r['ship_text']} |"
            )
        lines.append("")

    # Manufacturer mismatches — dedup by (input_mpn, source) since explode produces dupes
    mismatches: dict[tuple, dict] = {}
    for r in rows:
        if r["status"] != "ok":
            continue
        if r.get("mfr_match") is True:
            continue
        if not r.get("returned_mfr"):
            continue
        key = (r["input_mpn"], r["source"])
        if key not in mismatches:
            mismatches[key] = r
    lines.append("## Manufacturer mismatches (returned_mfr ≠ expected_mfr after normalization)")
    lines.append("")
    if not mismatches:
        lines.append("_None._")
    else:
        lines.append("| input_mpn | source | expected_mfr | returned_mfr |")
        lines.append("|---|---|---|---|")
        for r in list(mismatches.values())[:30]:
            lines.append(
                f"| `{r['input_mpn']}` | {r['source']} | "
                f"{r['expected_mfr']} | {r['returned_mfr']} |"
            )
        if len(mismatches) > 30:
            lines.append(f"| …and {len(mismatches) - 30} more (see `batch_index.csv`) |  |  |  |")
    lines.append("")

    # Skipped rows — collapse the high-cardinality "duplicate" reason into a
    # single count row; itemize the rare "missing/placeholder" rows fully.
    if skipped:
        dup_skipped = [s for s in skipped if "duplicate" in s.get("reason", "")]
        other_skipped = [s for s in skipped if "duplicate" not in s.get("reason", "")]
        lines.append("## Skipped xlsx rows")
        lines.append("")
        lines.append("| row | raw_mpn | raw_mfr | reason |")
        lines.append("|---|---|---|---|")
        for s in other_skipped:
            lines.append(
                f"| {s['row']} | `{s['raw_mpn']}` | {s['raw_mfr']} | {s['reason']} |"
            )
        if dup_skipped:
            lines.append(
                f"| — | — | — | {len(dup_skipped)} duplicate MPN rows "
                f"collapsed (dedup at load time) |"
            )
        lines.append("")

    lines.append("## Files in this batch folder")
    lines.append("")
    for name, what in [
        ("batch_summary.md", "this file"),
        ("batch_index.csv / .xlsx", "long form — one row per (MPN × source × warehouse)"),
        ("batch_index.json", "machine-readable long form (per-source raw records + chosen variant)"),
        ("batch_input.csv", "verbatim (MPN, expected_mfr) input rows"),
        ("failures.md", "non-ok (chip × source) pairs with error excerpt"),
        ("Test_<sanitized_mpn>_<SOURCE>/", "per-MPN run folder, one per source"),
    ]:
        lines.append(f"- `{name}` — {what}")
    lines.append("")

    (batch_dir / "batch_summary.md").write_text("\n".join(lines), encoding="utf-8")


# --- per-chip per-source worker (thread-safe; one instance runs per future) -


def _call_one_source(
    mpn: str,
    mfr: str,
    source: str,
    batch_dir: Path,
    source_min_interval: dict[str, float],
    source_last_call: dict[str, float],
    rate_limit_locks: dict[str, "threading.Lock"],
) -> dict:
    """Run one source for one chip. Returns a dict with:
       - source: str
       - index_rows: list[dict]   — warehouse rows (or 1 empty row on failure)
       - all_record: dict          — entry for batch_index.json
       - log_line: str             — caller prints this after the future returns

    Thread-safety notes:
      - Each source only touches its own key in `source_last_call`. The lock
        is defensive (in case future code submits multiple futures per source).
      - File writes go to a source-specific run_dir — no cross-thread file
        contention.
      - `os.environ` reads, `requests.post/get`, and JSON serialization are
        all thread-safe under CPython's GIL.
    """
    # Per-source rate-limit gate (Element14 only, in practice). Lock both
    # the read of last_call and the sleep, so concurrent threads on the same
    # source don't both see the same stale timestamp.
    min_gap = source_min_interval.get(source, 0.0)
    if min_gap > 0:
        with rate_limit_locks[source]:
            elapsed = time.time() - source_last_call[source]
            if elapsed < min_gap:
                time.sleep(min_gap - elapsed)
            source_last_call[source] = time.time()

    safe = _safe_folder(mpn)
    run_dir = batch_dir / f"Test_{safe}_{source}"
    run_dir.mkdir(parents=True, exist_ok=True)
    error: str | None = None
    bundle: dict | None = None
    t0 = time.time()
    try:
        bundle = SOURCE_RUNNERS[source](mpn, run_dir)
    except Exception as exc:
        error = f"{type(exc).__name__}: {exc}"
        traceback.print_exc()
    dt = time.time() - t0
    # Bump last_call AFTER completion as well, so the next call's gap is
    # measured from this call's end rather than its start.
    if min_gap > 0:
        with rate_limit_locks[source]:
            source_last_call[source] = time.time()

    rec = (bundle or {}).get("record") or {}
    chosen = (bundle or {}).get("chosen_extracted")
    status = rec.get("status") or ("exception" if error else "no_results")

    if status == "ok" and chosen:
        # Same-MPN multi-SKU expansion (Element14 only as of 2026-05). When
        # `expanded_extracteds` has >1 entry, iterate each — they share the
        # same returned_mpn but differ in vendor_sku + packaging_option.
        expanded = bundle.get("expanded_extracteds") if bundle else None
        if expanded and len(expanded) > 1:
            new_rows = []
            for variant_ex in expanded:
                new_rows.extend(iter_warehouse_rows(
                    mpn, mfr, source, bundle, run_dir, ex_override=variant_ex
                ))
            # Renumber warehouse_idx globally for this (chip × source) group.
            for i, r in enumerate(new_rows, start=1):
                r["warehouse_idx"] = i
        else:
            new_rows = list(iter_warehouse_rows(mpn, mfr, source, bundle, run_dir))
        if not new_rows:
            new_rows = [make_empty_source_row(
                mpn, mfr, source, "no_results",
                "variant returned but stock_breakdown empty", run_dir
            )]
        summary_qty = sum((r["stockpool_qty"] or 0) for r in new_rows) or 0
        log_line = (f"      {source}: ok  warehouses={len(new_rows)}  "
                    f"sum_qty={summary_qty}  ({dt:.2f} s)")
    else:
        new_rows = [make_empty_source_row(mpn, mfr, source, status, error or "", run_dir)]
        log_line = f"      {source}: {status}  ({dt:.2f} s)"

    all_record = {
        "source": _source_display(source),
        "input_mpn": mpn,
        "expected_mfr": mfr,
        "elapsed_sec": round(dt, 3),
        "record": rec,
        "extracted_best": chosen,
        "error": error,
    }
    return {
        "source": source,  # short code — used as dict key by main loop
        "index_rows": new_rows,
        "all_record": all_record,
        "log_line": log_line,
    }


# --- main -------------------------------------------------------------------


def main(argv: list[str]) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX,
                        help=f"Path to the chip-list xlsx (default: {DEFAULT_XLSX})")
    parser.add_argument("--mpns", type=str, default=None,
                        help="Semicolon-separated MPN list — overrides --xlsx entirely. "
                             "Optional `MPN:expected_mfr` syntax per entry: "
                             "'STM32G030F6P6:STM;BT168GW,115:WEEN'. NB: MPNs containing "
                             "`:` (e.g. typo'd `BTA206X-800CT:127`) will be wrongly "
                             "chopped — use --mpns-file with tab separator for those.")
    parser.add_argument("--mpns-file", type=str, default=None,
                        help="Tab-separated file: one MPN per line, format 'MPN<TAB>MFR' "
                             "(MFR optional). Lines starting with `#` are comments. "
                             "Overrides --xlsx; takes precedence over --mpns.")
    parser.add_argument("--limit", type=int, default=None,
                        help="Process only the first N valid MPNs (dry-run aid).")
    parser.add_argument("--only", action="append", choices=SOURCES_ALL, default=None,
                        help="Run only the given source(s). Repeat to whitelist multiple. "
                             "Default: all four.")
    parser.add_argument("--throttle", type=float, default=THROTTLE_SECONDS,
                        help="Inter-chip pause (seconds). Within a chip the four "
                             "sources run in parallel; per-source rate limits are "
                             "enforced separately.")
    parser.add_argument("--max-workers", type=int, default=None,
                        help="Max concurrent sources per chip (default = number "
                             "of sources being run). Pass 1 to force serial mode "
                             "for debugging.")
    parser.add_argument("--env", choices=("test", "prod"), default="test",
                        help="Output root: 'test' → test/api/ (default), 'prod' → production/api/.")
    args = parser.parse_args(argv[1:])

    load_dotenv(ENV_PATH)
    sources_to_run = list(args.only) if args.only else list(SOURCES_ALL)

    # Credential preflight — fail fast if any requested source is missing keys.
    creds_missing: list[str] = []
    if "MOUSER" in sources_to_run and not os.environ.get("MOUSER_API_KEY"):
        creds_missing.append("MOUSER_API_KEY")
    if "DIGIKEY" in sources_to_run and not (
        os.environ.get("DIGIKEY_CLIENT_ID") and os.environ.get("DIGIKEY_CLIENT_SECRET")
    ):
        creds_missing.append("DIGIKEY_CLIENT_ID / DIGIKEY_CLIENT_SECRET")
    if "ELEMENT14" in sources_to_run and not os.environ.get("ELEMENT14_API_KEY"):
        creds_missing.append("ELEMENT14_API_KEY")
    if "ARROW" in sources_to_run and not (
        os.environ.get("ARROW_LOGIN") and os.environ.get("ARROW_API_KEY")
    ):
        creds_missing.append("ARROW_LOGIN / ARROW_API_KEY")
    if "LCSC" in sources_to_run and not (
        os.environ.get("lcsc_AppID")
        and os.environ.get("lcsc_AccessKey")
        and os.environ.get("lcsc_SecretKey")
    ):
        creds_missing.append("lcsc_AppID / lcsc_AccessKey / lcsc_SecretKey")
    if creds_missing:
        print(f"ERROR: missing in api/.env — {', '.join(creds_missing)}", file=sys.stderr)
        return 2

    # Chip-list source: --mpns-file > --mpns > --xlsx (mirrors scraper batch driver).
    if args.mpns_file:
        chips = []
        skipped = []
        for i, line in enumerate(Path(args.mpns_file).read_text(encoding="utf-8").splitlines()):
            line = line.rstrip("\r\n")
            if not line.strip() or line.lstrip().startswith("#"):
                continue
            if "\t" in line:
                mpn_s, mfr_s = line.split("\t", 1)
            else:
                mpn_s, mfr_s = line, ""
            chips.append({"row": i + 1, "input_mpn": mpn_s.strip(), "expected_mfr": mfr_s.strip()})
        print(f"Loaded {len(chips)} chip rows from --mpns-file (xlsx ignored)")
    elif args.mpns:
        chips = []
        skipped = []
        for i, raw in enumerate(args.mpns.split(";")):
            entry = raw.strip()
            if not entry:
                continue
            if ":" in entry:
                mpn_s, mfr_s = entry.split(":", 1)
                mpn_s, mfr_s = mpn_s.strip(), mfr_s.strip()
            else:
                mpn_s, mfr_s = entry, ""
            chips.append({"row": i + 1, "input_mpn": mpn_s, "expected_mfr": mfr_s})
        print(f"Loaded {len(chips)} chip rows from --mpns flag (xlsx ignored)")
    else:
        chips, skipped = load_chip_list(args.xlsx)
        print(f"Loaded {len(chips)} chip rows ({len(skipped)} skipped) from {args.xlsx.name}")
    if args.limit:
        chips = chips[: args.limit]
    print(f"Sources: {', '.join(sources_to_run)}")

    now = datetime.now()
    batch_name = f"BatchTest_{now.strftime('%Y%m%d')}_{now.strftime('%H_%M_%S')}"
    api_root = ENV_ROOTS[args.env] / "api"
    batch_dir = api_root / batch_name
    batch_dir.mkdir(parents=True, exist_ok=True)
    print(f"Env: {args.env}  →  Batch folder: {batch_dir}")

    write_batch_input_csv(chips, batch_dir / "batch_input.csv")

    started = datetime.now(timezone.utc)
    index_rows: list[dict] = []
    all_records: list[dict] = []  # for batch_index.json (full detail)

    # Per-source minimum-interval table (resolved AFTER .env is loaded)
    source_min_interval = dict(SOURCE_MIN_INTERVAL_DEFAULT)
    source_min_interval["ELEMENT14"] = _element14_min_interval()
    source_last_call: dict[str, float] = {src: 0.0 for src in SOURCES_ALL}
    rate_limit_locks: dict[str, threading.Lock] = {src: threading.Lock() for src in SOURCES_ALL}
    print_lock = threading.Lock()

    # Workers default to one per source (full parallelism within a chip).
    # User can pass --max-workers 1 to force serial for debugging.
    if args.max_workers is None:
        max_workers = len(sources_to_run)
    else:
        max_workers = max(1, min(args.max_workers, len(sources_to_run)))
    print(f"Parallelism: {max_workers} source(s) concurrent per chip")

    for i, chip in enumerate(chips, 1):
        mpn = chip["input_mpn"]
        mfr = chip["expected_mfr"]
        with print_lock:
            print(f"[{i:>3}/{len(chips)}] {mpn}  (expected {mfr})")

        results_by_source: dict[str, dict] = {}
        with ThreadPoolExecutor(max_workers=max_workers) as ex:
            futures = {
                ex.submit(
                    _call_one_source, mpn, mfr, src, batch_dir,
                    source_min_interval, source_last_call, rate_limit_locks,
                ): src
                for src in sources_to_run
            }
            for fut in as_completed(futures):
                src = futures[fut]
                try:
                    res = fut.result()
                except Exception as exc:
                    # Defensive: any exception escaping _call_one_source
                    # (shouldn't happen — it catches its own).
                    with print_lock:
                        print(f"      {src}: WORKER_EXCEPTION {type(exc).__name__}: {exc}")
                    continue
                with print_lock:
                    print(res["log_line"])
                results_by_source[res["source"]] = res

        # Collect into the global lists in deterministic (sources_to_run)
        # order so batch_index row ordering is stable across runs.
        for src in sources_to_run:
            r = results_by_source.get(src)
            if r is None:
                continue
            index_rows.extend(r["index_rows"])
            all_records.append(r["all_record"])

        if args.throttle > 0 and i < len(chips):
            time.sleep(args.throttle)

    finished = datetime.now(timezone.utc)

    # Write all output files
    write_csv(index_rows, INDEX_COLUMNS, batch_dir / "batch_index.csv")
    write_xlsx(index_rows, INDEX_COLUMNS, batch_dir / "batch_index.xlsx", "batch_index")
    (batch_dir / "batch_index.json").write_text(
        json.dumps(all_records, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    write_failures(index_rows, batch_dir / "failures.md")
    write_summary_md(
        chips,
        index_rows,
        skipped,
        batch_dir,
        started.astimezone(),
        finished.astimezone(),
        sources_to_run,
    )

    print(f"\nDone. {len(index_rows)} warehouse rows across {len(chips)} chips × "
          f"{len(sources_to_run)} sources.")
    print(f"Wrote: {batch_dir}")

    # Refresh api/README.md status block — best-effort, PROD RUNS ONLY.
    # The committed README is a prod-facing artifact; test sweeps must not
    # clobber it. _update_readme_status.py scans production/api/ for the latest
    # batch, so gating here keeps the snapshot in sync with what the scanner reads.
    regen = PROJECT_ROOT / "api" / "scripts" / "_update_readme_status.py"
    if args.env == "prod" and regen.exists():
        try:
            import subprocess
            subprocess.run(
                [sys.executable, str(regen)],
                timeout=10,
                check=False,
                capture_output=True,
            )
            print(f"Refreshed: api/README.md status block")
        except (subprocess.TimeoutExpired, OSError):
            pass
    elif args.env != "prod":
        print("Skipped api/README.md refresh (test run — README tracks prod only)")

    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
